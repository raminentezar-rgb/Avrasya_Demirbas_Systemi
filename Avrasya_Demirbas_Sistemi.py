# Avrasya_Demirbas_Sistemi.py

import sys
import os
from customtkinter import *
import tkinter.ttk as ttk
from tkcalendar import DateEntry
from openpyxl import Workbook, load_workbook
from fpdf import FPDF
import csv, re, shutil
from uuid import uuid4
from time import strftime
import matplotlib.pyplot as plt
from collections import Counter, defaultdict
from PIL import Image, ImageTk
import pandas as pd
from tkinter import filedialog, messagebox
import barcode
from barcode.writer import ImageWriter
import qrcode




def resource_path(relative_path):
    """Get absolute path to resource, works for dev and for PyInstaller"""
    try:
        # PyInstaller creates a temp folder and stores path in _MEIPASS
        base_path = sys._MEIPASS
    except Exception:
        base_path = os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base_path, relative_path)

# تنظیمات مسیرها
if getattr(sys, 'frozen', False):
    # Running as compiled exe
    base_path_exe = os.path.dirname(sys.executable)
    EXCEL_FILE = os.path.join(base_path_exe, "inventory.xlsx")
    BARCODE_DIR = os.path.join(base_path_exe, "barcodes")
    QRCODE_DIR = os.path.join(os.path.dirname(sys.executable), "qrcodes")
    IMAGES_DIR = os.path.join(base_path_exe, "images")
    LOGO_FILE = resource_path("LOGO1.png")
else:
    # Development mode
    EXCEL_FILE = "inventory.xlsx"
    BARCODE_DIR = "barcodes"
    QRCODE_DIR = "qrcodes"
    IMAGES_DIR = "images"
    LOGO_FILE = "LOGO1.png"

HEADERS = ["ID","Code","Internal_Code","Inventory_Item","Campus","Floor","Room","Category","Status","Date","Color","Turkish_Name","Image_Path"]

# ---------- Create Directories ----------
def create_directories():
    """ایجاد پوشه‌های مورد نیاز"""
    os.makedirs(BARCODE_DIR, exist_ok=True)
    os.makedirs(QRCODE_DIR, exist_ok=True)
    os.makedirs(IMAGES_DIR, exist_ok=True)

# ---------- Create/Check Excel File ----------
def init_workbook():
    """ایجاد یا بررسی فایل اکسل"""
    create_directories()
    
    if not os.path.exists(EXCEL_FILE):
        # ایجاد فایل جدید
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)
        print(f"فایل جدید ایجاد شد: {EXCEL_FILE}")
    else:
        # بررسی ساختار فایل موجود
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        
        # بررسی هدرها
        first_row = [cell.value for cell in ws[1]]
        if first_row != HEADERS:
            print("هدرهای فایل قدیمی است. در حال بروزرسانی...")
            
            # ذخیره داده‌های قدیمی
            old_data = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                old_data.append(row)
            
            # ایجاد فایل جدید با هدرهای صحیح
            wb = Workbook()
            ws = wb.active
            ws.append(HEADERS)
            
            # اضافه کردن داده‌های قدیمی با ID جدید
            for row in old_data:
                row_list = list(row)
                
                # اگر row دارای ID است، آن را حذف کن (ستون اول)
                # چون ما ID جدید اضافه خواهیم کرد
                if len(row_list) > 0 and row_list[0]:  # اگر ID دارد
                    # فرض می‌کنیم اولین ستون ID است، آن را نگه می‌داریم
                    existing_id = row_list[0]
                    data_without_id = row_list[1:]
                else:
                    existing_id = str(uuid4())
                    data_without_id = row_list
                
                # اگر داده‌ها کمتر از تعداد مورد نیاز هستند، خالی اضافه کن
                # HEADERS has 13 items: ID + 11 data fields + Image_Path
                # So we need 12 items in data_without_id
                while len(data_without_id) < len(HEADERS) - 1:
                    data_without_id.append("")
                
                # اکنون new_row = [ID] + 12 data fields = 13 items
                new_row = [existing_id] + data_without_id[:len(HEADERS)-1]
                ws.append(new_row)
            
            wb.save(EXCEL_FILE)
            print("فایل اکسل با ساختار جدید ذخیره شد.")
        else:
            print("ساختار فایل اکسل صحیح است.")
    
    return True

# بقیه کدها...

# ---------- Barcode Functions ----------
# def generate_barcode(item_code, item_name):
#     """Generate barcode for an item"""
#     try:
#         # اطمینان از اینکه item_code رشته است
#         if not isinstance(item_code, str):
#             item_code = str(item_code)
        
#         # حذف فاصله‌ها و کاراکترهای غیرمجاز
#         item_code = item_code.strip()
        
#         # ایجاد پوشه اگر وجود ندارد
#         os.makedirs(BARCODE_DIR, exist_ok=True)
        
#         # Create CODE128 barcode
#         barcode_class = barcode.get_barcode_class('code128')
#         barcode_obj = barcode_class(item_code, writer=ImageWriter())
        
#         # تنظیمات بارکد
#         options = {
#             'write_text': True,
#             'text_distance': 2,
#             'font_size': 12,
#             'module_height': 15.0,
#             'quiet_zone': 6.5
#         }
        
#         # مسیر کامل فایل (بدون پسوند)
#         filename = f"barcode_{item_code}"
#         full_path = os.path.join(BARCODE_DIR, filename)
        
#         # ذخیره بارکد
#         saved_path = barcode_obj.save(full_path, options=options)
        
#         print(f"DEBUG: Barcode generated for '{item_code}' at: {saved_path}")
#         return saved_path
        
#     except Exception as e:
#         print(f"Barcode generation error for code '{item_code}': {e}")
#         import traceback
#         traceback.print_exc()  # نمایش کامل traceback برای دیباگ
#         return None
    

def generate_qrcode(item_data, item_code):
    """Generate QR code for an item with complete information"""
    try:
        # Create QR code with COMPLETE item information
        qr = qrcode.QRCode(
            version=2,  # Increased version for more data
            error_correction=qrcode.constants.ERROR_CORRECT_M,
            box_size=8,
            border=4,
        )
        
        # Create detailed information string
        qr_data = f""" AVRASYA ÜNİVERSİTESİ - DEMİRBAŞ SİSTEMİ 

TEMEL BİLGİLER:
├─ Demirbaş Nomarası: {item_code}
├─ Dahili Kod: {item_data.get('internal_code', '')}
├─ Kullanan Personel: {item_data['turkish_name']}
├─ Ad: {item_data.get('inventory_item', '')}

YERLEŞİM BİLGİSİ:
├─ Yerleşke: {item_data['campus']}
├─ Kat: {item_data.get('floor', '')}
├─ Oda: {item_data['room']}

KATEGORİ ve DURUM:
├─ Kategori: {item_data['category']}
├─ Durum: {item_data['status']}
├─ Renk: {item_data.get('color', 'Belirtilmemiş')}

KAYIT BİLGİSİ:
├─ Kayıt Tarihi: {item_data.get('date', '')}
├─ Benzersiz ID: {item_data.get('id', '')}

AÇIKLAMA:
Avrasya Üniversitesi demirbaş sistemine kayıtlı demirbaş malzemedir.
Lütfen bu malzemenin yerini değiştirmeyiniz veya başka birime taşımayınız.

 İLETİŞİM:
Demirbaş Yönetimi - Avrasya Üniversitesi
"""

        qr.add_data(qr_data)
        qr.make(fit=True)
        
        qr_img = qr.make_image(fill_color="black", back_color="white")
        qr_path = os.path.join(QRCODE_DIR, f"qrcode_{item_code}.png")
        qr_img.save(qr_path)
        
        return qr_path
    except Exception as e:
        print(f"QR code generation error: {e}")
        return None

# def generate_detailed_barcode(item_data, item_code):
#     """Generate barcode with detailed information encoded"""
#     try:
#         # Create a detailed data string for barcode (limited capacity)
#         barcode_data = f"{item_code}|{item_data['turkish_name'][:20]}|{item_data['campus']}|{item_data['room']}"
        
#         barcode_class = barcode.get_barcode_class('code128')
#         barcode_obj = barcode_class(barcode_data, writer=ImageWriter())
        
#         filename = os.path.join(BARCODE_DIR, f"detailed_barcode_{item_code}")
#         barcode_path = barcode_obj.save(filename, options={
#             'write_text': True,
#             'text_distance': 2,
#             'font_size': 10,
#             'module_height': 12.0,
#             'quiet_zone': 4.5
#         })
        
#         return barcode_path
#     except Exception as e:
#         print(f"Detailed barcode generation error: {e}")
#         return None

# def show_barcode():
#     """Show barcode for selected item"""
#     sel = tree.selection()
#     if not sel:
#         info_var.set("Lütfen bir öğe seçin.")
#         return
    
#     vals = tree.item(sel[0])['values']
#     item_code = vals[1]
#     item_name = vals[11] if len(vals) > 11 else vals[3]
    
#     # دیباگ: نمایش نوع داده‌ها
#     print(f"DEBUG: Item code type: {type(item_code)}, value: '{item_code}'")
#     print(f"DEBUG: Item name: '{item_name}'")
    
#     # Generate barcode
#     barcode_path = generate_barcode(item_code, item_name)
    
#     if barcode_path:
#         print(f"DEBUG: Barcode path returned: {barcode_path}")
        
#         # Show barcode in new window
#         barcode_win = CTkToplevel(w)
#         barcode_win.title(f"Barkod - {item_code}")
#         barcode_win.geometry("500x400")
        
#         try:
#             # استفاده از CTkImage به جای ImageTk.PhotoImage
#             from customtkinter import CTkImage
            
#             # بارگذاری تصویر
#             barcode_img = Image.open(barcode_path)
#             barcode_img = barcode_img.resize((400, 150), Image.LANCZOS)
            
#             # تبدیل به CTkImage
#             barcode_ctk_image = CTkImage(light_image=barcode_img, size=(400, 150))
            
#             CTkLabel(barcode_win, text=f"BARKOD - {item_code}", 
#                      font=("Arial", 16, "bold")).pack(pady=10)
#             CTkLabel(barcode_win, text=item_name, 
#                      font=("Arial", 12)).pack(pady=5)
#             CTkLabel(barcode_win, text="Sadece demirbaş numarası içerir", 
#                      font=("Arial", 10), text_color="gray").pack(pady=2)
            
#             # استفاده از CTkLabel با CTkImage
#             img_label = CTkLabel(barcode_win, image=barcode_ctk_image, text="")
#             img_label.pack(pady=10)
            
#         except Exception as e:
#             error_msg = f"Barkod yüklenemedi: {str(e)}"
#             print(f"Error loading barcode: {error_msg}")
#             CTkLabel(barcode_win, text=error_msg, 
#                      font=("Arial", 12)).pack(pady=20)
        
#         # Print button
#         CTkButton(barcode_win, text="Barkodu Yazdır", 
#                  command=lambda: print_barcode(barcode_path, item_code, item_name),
#                  width=140, height=35).pack(pady=10)
#     else:
#         info_var.set("Barkod oluşturulamadı.")
#         print("DEBUG: Barcode generation failed")



def show_qrcode():
    """Show QR code for selected item with complete information"""
    sel = tree.selection()
    if not sel:
        info_var.set("Lütfen bir öğe seçin.")
        return
    
    vals = tree.item(sel[0])['values']
    item_code = vals[1]
    
    # Prepare COMPLETE item data
    item_data = {
        'id': vals[0],
        'internal_code': vals[2],
        'inventory_item': vals[3],
        'turkish_name': vals[11] if len(vals) > 11 else vals[3],
        'campus': vals[4],
        'floor': vals[5],
        'room': vals[6],
        'category': vals[7],
        'status': vals[8],
        'date': vals[9],
        'color': vals[10] if len(vals) > 10 else ""
    }
    
    # Generate QR code
    qr_path = generate_qrcode(item_data, item_code)
    
    if qr_path:
        # Show QR code in new window
        qr_win = CTkToplevel(w)
        qr_win.title(f"QR Kod - {item_code}")
        qr_win.geometry("400x550")
        
        # Load and display QR code image using CTkImage
        try:
            qr_img = Image.open(qr_path)
            qr_img = qr_img.resize((280, 280), Image.LANCZOS)
            
            # تبدیل به CTkImage
            qr_ctk_image = CTkImage(light_image=qr_img, dark_image=qr_img, size=(280, 280))
            
            CTkLabel(qr_win, text="QR KOD - TÜM BİLGİLER", font=("Arial", 16, "bold")).pack(pady=10)
            
            # استفاده از CTkLabel با CTkImage
            img_label = CTkLabel(qr_win, image=qr_ctk_image, text="")
            img_label.pack(pady=10)
            
            # Item information display
            info_frame = CTkFrame(qr_win)
            info_frame.pack(pady=10, padx=20, fill="x")
            
            CTkLabel(info_frame, text=f"Demirbaş No: {item_code}", font=("Arial", 12, "bold")).pack(anchor="w", pady=2)
            CTkLabel(info_frame, text=f"Malzeme: {item_data['turkish_name']}", font=("Arial", 11)).pack(anchor="w", pady=1)
            CTkLabel(info_frame, text=f"Yer: {item_data['campus']} - {item_data['room']}", font=("Arial", 11)).pack(anchor="w", pady=1)
            CTkLabel(info_frame, text=f"Durum: {item_data['status']}", font=("Arial", 11)).pack(anchor="w", pady=1)
            CTkLabel(info_frame, text="Tüm bilgiler QR kod içinde mevcut", font=("Arial", 10), text_color="green").pack(anchor="w", pady=5)
            
        except Exception as e:
            CTkLabel(qr_win, text=f"QR kod yüklenemedi: {str(e)}", font=("Arial", 12)).pack(pady=20)
        
        # Print button
        CTkButton(qr_win, text="QR Kod Yazdır", 
                 command=lambda: print_qrcode(qr_path, item_code, item_data),
                 width=140, height=35).pack(pady=10)
    else:
        info_var.set("QR kod oluşturulamadı.")
# def show_detailed_barcode():
#     """Show detailed barcode with more information"""
#     sel = tree.selection()
#     if not sel:
#         info_var.set("Lütfen bir öğe seçin.")
#         return
    
#     vals = tree.item(sel[0])['values']
#     item_code = vals[1]
    
#     item_data = {
#         'turkish_name': vals[11] if len(vals) > 11 else vals[3],
#         'campus': vals[4],
#         'room': vals[6]
#     }
    
#     # Generate detailed barcode
#     barcode_path = generate_detailed_barcode(item_data, item_code)
    
#     if barcode_path:
#         barcode_win = CTkToplevel(w)
#         barcode_win.title(f"Detaylı Barkod - {item_code}")
#         barcode_win.geometry("500x400")
        
#         barcode_img_path = barcode_path + ".png" if not barcode_path.endswith('.png') else barcode_path
#         try:
#             barcode_img = Image.open(barcode_img_path)
#             barcode_img = barcode_img.resize((450, 180), Image.LANCZOS)
#             barcode_photo = ImageTk.PhotoImage(barcode_img)
            
#             CTkLabel(barcode_win, text="DETAYLI BARKOD", font=("Arial", 16, "bold")).pack(pady=10)
#             CTkLabel(barcode_win, text=f"Demirbaş No: {item_code}", font=("Arial", 12)).pack(pady=2)
#             CTkLabel(barcode_win, text=f"Malzeme: {item_data['turkish_name']}", font=("Arial", 11)).pack(pady=1)
#             CTkLabel(barcode_win, text="Sınırlı bilgi içerir (Demirbaş no, isim, yer)", font=("Arial", 10), text_color="orange").pack(pady=2)
            
#             from tkinter import Label
#             img_label = Label(barcode_win, image=barcode_photo)
#             img_label.image = barcode_photo
#             img_label.pack(pady=10)
            
#         except Exception as e:
#             CTkLabel(barcode_win, text=f"Barkod yüklenemedi: {str(e)}", font=("Arial", 12)).pack(pady=20)
        
#         CTkButton(barcode_win, text="Yazdır", 
#                  command=lambda: print_barcode(barcode_path, item_code, item_data['turkish_name']),
#                  width=120).pack(pady=10)
#     else:
#         info_var.set("Detaylı barkod oluşturulamadı.")

# def print_barcode(barcode_path, item_code, item_name):
#     """Print barcode"""
#     try:
#         # اطمینان از رشته بودن کد
#         if not isinstance(item_code, str):
#             item_code = str(item_code)
            
#         pdf = FPDF()
#         pdf.add_page()
        
#         # Title
#         pdf.set_font("Arial", 'B', 16)
#         pdf.cell(200, 10, "AVRASYA ÜNİVERSİTESİ", ln=True, align='C')
#         pdf.ln(5)
        
#         # Item info
#         pdf.set_font("Arial", '', 12)
#         pdf.cell(200, 8, f"Demirbaş No: {item_code}", ln=True, align='C')
#         pdf.cell(200, 8, f"Malzeme: {item_name}", ln=True, align='C')
#         pdf.cell(200, 8, "Sadece Demirbaş numarası içerir", ln=True, align='C')
#         pdf.ln(10)
        
#         # Barcode image - اضافه کردن پسوند اگر لازم باشد
#         if not barcode_path.endswith('.png'):
#             barcode_path = barcode_path + '.png'
            
#         if os.path.exists(barcode_path):
#             pdf.image(barcode_path, x=40, y=60, w=130)
#         else:
#             # اگر فایل با پسوند نباشد، بدون پسوند امتحان کن
#             if os.path.exists(barcode_path.replace('.png', '')):
#                 pdf.image(barcode_path.replace('.png', ''), x=40, y=60, w=130)
#             else:
#                 pdf.cell(200, 10, "Barkod resmi bulunamadı", ln=True, align='C')
        
#         # Footer
#         pdf.set_y(-30)
#         pdf.set_font("Arial", 'I', 10)
#         pdf.cell(200, 10, f"Oluşturulma: {strftime('%Y-%m-%d %H:%M')}", ln=True, align='C')
        
#         filename = f"barcode_print_{item_code}.pdf"
#         pdf.output(filename)
        
#         info_var.set(f"Barkod yazdırma dosyası oluşturuldu: {filename}")
#         messagebox.showinfo("Başarılı", f"Barkod yazdırma dosyası oluşturuldu:\n{filename}")
        
#     except Exception as e:
#         error_msg = f"Yazdırma hatası: {str(e)}"
#         print(f"Print error: {error_msg}")
#         messagebox.showerror("Hata", error_msg)

def print_qrcode(qr_path, item_code, item_data):
    """Print QR code with information exactly like the image"""
    try:
        pdf = FPDF()
        pdf.add_page()
        
        # Add university logo at top center
        if os.path.exists(LOGO_FILE):
            pdf.image(LOGO_FILE, x=75, y=10, w=60)
            pdf.set_y(40)  # Start below logo
        else:
            pdf.set_y(10)
        
        # Title
        pdf.set_font("Arial", 'B', 18)
        pdf.set_text_color(0, 51, 102)  # Dark blue
        pdf.cell(200, 10, "AVRASYA ÜNİVERSİTESİ - DEMİRBAŞ SİSTEMİ", ln=True, align='C')
        pdf.ln(8)
        
        # Separator line
        pdf.set_draw_color(0, 51, 102)
        pdf.line(20, pdf.get_y(), 190, pdf.get_y())
        pdf.ln(10)
        
        # Main content in two columns
        current_y = pdf.get_y()
        
        # Left column - Item information
        pdf.set_font("Arial", 'B', 14)
        pdf.set_text_color(0, 0, 0)
        pdf.set_xy(20, current_y)
        pdf.cell(80, 8, " TEMEL BİLGİLER:", ln=True)
        pdf.set_font("Arial", '', 11)
        pdf.set_x(20)
        pdf.cell(80, 7, f"  • Demirbaş No: {item_code}", ln=True)
        pdf.set_x(20)
        pdf.cell(80, 7, f"  • Dahili Kod: {item_data.get('internal_code', '')}", ln=True)
        pdf.set_x(20)
        pdf.cell(80, 7, f"  • Malzeme: {item_data['turkish_name']}", ln=True)
        pdf.set_x(20)
        pdf.cell(80, 7, f"  • İngilizce Ad: {item_data.get('inventory_item', '')}", ln=True)
        
        pdf.ln(5)
        
        # Location information
        pdf.set_font("Arial", 'B', 14)
        pdf.set_x(20)
        pdf.cell(80, 8, " YERLEŞİM BİLGİSİ:", ln=True)
        pdf.set_font("Arial", '', 11)
        pdf.set_x(20)
        pdf.cell(80, 7, f"  • Yerleşke: {item_data['campus']}", ln=True)
        pdf.set_x(20)
        pdf.cell(80, 7, f"  • Kat: {item_data.get('floor', '')}", ln=True)
        pdf.set_x(20)
        pdf.cell(80, 7, f"  • Oda: {item_data['room']}", ln=True)
        
        # Right column - Category and status
        pdf.set_font("Arial", 'B', 14)
        pdf.set_xy(110, current_y)
        pdf.cell(80, 8, " KATEGORİ ve DURUM:", ln=True)
        pdf.set_font("Arial", '', 11)
        pdf.set_x(110)
        pdf.cell(80, 7, f"  • Kategori: {item_data['category']}", ln=True)
        pdf.set_x(110)
        pdf.cell(80, 7, f"  • Durum: {item_data['status']}", ln=True)
        pdf.set_x(110)
        pdf.cell(80, 7, f"  • Renk: {item_data.get('color', 'Belirtilmemiş')}", ln=True)
        
        pdf.ln(5)
        
        # Registration info
        pdf.set_font("Arial", 'B', 14)
        pdf.set_x(110)
        pdf.cell(80, 8, " KAYIT BİLGİSİ:", ln=True)
        pdf.set_font("Arial", '', 11)
        pdf.set_x(110)
        pdf.cell(80, 7, f"  • Kayıt Tarihi: {item_data.get('date', '')}", ln=True)
        pdf.set_x(110)
        pdf.cell(80, 7, f"  • Benzersiz ID: {item_data.get('id', '')}", ln=True)
        
        pdf.ln(15)
        
        # Description
        pdf.set_font("Arial", 'B', 14)
        pdf.set_x(20)
        pdf.cell(180, 8, " AÇIKLAMA:", ln=True)
        pdf.set_font("Arial", '', 11)
        pdf.set_x(20)
        description = "Avrasya Üniversitesi demirbaş sistemine kayıtlı demirbaş malzemedir. Lütfen bu malzemenin yerini değiştirmeyiniz veya başka birime taşımayınız."
        pdf.multi_cell(170, 7, description)
        
        pdf.ln(10)
        
        # QR code at bottom center
        pdf.image(qr_path, x=60, y=pdf.get_y(), w=80)
        
        # Footer
        pdf.set_y(-30)
        pdf.set_font("Arial", 'I', 10)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(200, 6, "Demirbaş Yönetimi - Avrasya Üniversitesi", ln=True, align='C')
        pdf.cell(200, 6, f"Belge oluşturulma: {strftime('%Y-%m-%d %H:%M')}", ln=True, align='C')
        
        filename = f"Demirbaş_belgesi_{item_code}.pdf"
        pdf.output(filename)
        
        info_var.set(f"Demirbaş belgesi oluşturuldu: {filename}")
        messagebox.showinfo("Başarılı", f"Demirbaş bilgi belgesi oluşturuldu:\n{filename}")
        
    except Exception as e:
        messagebox.showerror("Hata", f"Yazdırma hatası: {str(e)}")

# ---------- Image Upload and Display Functions ----------
def upload_item_image():
    """Upload an image for the selected item"""
    sel = tree.selection()
    if not sel:
        info_var.set("Lütfen bir öğe seçin.")
        messagebox.showwarning("Uyarı", "Lütfen önce bir öğe seçin!")
        return
    
    # Get item details
    vals = tree.item(sel[0])['values']
    item_id = vals[0]
    item_code = vals[1]
    item_name = vals[11] if len(vals) > 11 else vals[3]
    
    # Open file dialog to select image
    file_path = filedialog.askopenfilename(
        title="Görsel Seçin / Resim Seçin",
        filetypes=[
            ("Resim Dosyaları", "*.png *.jpg *.jpeg *.gif *.bmp"),
            ("Tüm Dosyalar", "*.*")
        ]
    )
    
    if not file_path:
        return
    
    try:
        # Create unique filename
        file_extension = os.path.splitext(file_path)[1]
        new_filename = f"item_{item_code}{file_extension}"
        destination_path = os.path.join(IMAGES_DIR, new_filename)
        
        # Copy image to images directory
        shutil.copy(file_path, destination_path)
        
        # Update database with image path
        update_item_image_path(item_id, destination_path)
        
        # Reload tree to show updated data
        load_to_tree()
        
        # Show success message and display image
        info_var.set(f"Görsel yüklendi: {new_filename}")
        messagebox.showinfo("Başarılı", f"Görsel başarıyla yüklendi!\n{new_filename}")
        
        # Automatically show the uploaded image
        show_item_image()
        
    except Exception as e:
        error_msg = f"Görsel yükleme hatası: {str(e)}"
        info_var.set(error_msg)
        messagebox.showerror("Hata", error_msg)

def update_item_image_path(item_id, image_path):
    """Update the image path for an item in the database"""
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(item_id):
            # Image_Path is the last column (index 12 in HEADERS)
            row[12].value = image_path
            break
    
    wb.save(EXCEL_FILE)

def show_item_image():
    """Display the image for the selected item"""
    sel = tree.selection()
    if not sel:
        info_var.set("Lütfen bir öğe seçin.")
        messagebox.showwarning("Uyarı", "Lütfen önce bir öğe seçin!")
        return
    
    vals = tree.item(sel[0])['values']
    item_code = vals[1]
    item_name = vals[11] if len(vals) > 11 else vals[3]
    image_path = vals[12] if len(vals) > 12 and vals[12] else None
    
    if not image_path or not os.path.exists(image_path):
        messagebox.showinfo("Bilgi", "Bu öğe için görsel bulunamadı.\nLütfen önce görsel yükleyin.")
        return
    
    try:
        # Create image window
        img_win = CTkToplevel(w)
        img_win.title(f"Görsel - {item_code}")
        img_win.geometry("600x700")
        
        # Title
        CTkLabel(img_win, text="DEMİRBAŞ GÖRSELİ", 
                 font=("Arial", 18, "bold")).pack(pady=10)
        
        # Item info
        info_frame = CTkFrame(img_win)
        info_frame.pack(pady=10, padx=20, fill="x")
        
        CTkLabel(info_frame, text=f"Demirbaş No: {item_code}", 
                 font=("Arial", 13, "bold")).pack(anchor="w", pady=3)
        CTkLabel(info_frame, text=f"Malzeme: {item_name}", 
                 font=("Arial", 12)).pack(anchor="w", pady=2)
        
        # Load and display image
        img = Image.open(image_path)
        
        # Resize image to fit window while maintaining aspect ratio
        img.thumbnail((550, 450), Image.LANCZOS)
        
        # Convert to CTkImage
        ctk_image = CTkImage(light_image=img, dark_image=img, 
                            size=(img.width, img.height))
        
        # Display image
        img_label = CTkLabel(img_win, image=ctk_image, text="")
        img_label.image = ctk_image  # Keep a reference
        img_label.pack(pady=15)
        
        # Path info
        CTkLabel(img_win, text=f"Dosya: {os.path.basename(image_path)}", 
                 font=("Arial", 10), text_color="gray").pack(pady=5)
        
        # Delete image button
        def delete_image():
            confirm = messagebox.askyesno("Onay", "Bu görseli silmek istediğinizden emin misiniz?")
            if confirm:
                try:
                    # Delete file
                    if os.path.exists(image_path):
                        os.remove(image_path)
                    
                    # Update database
                    update_item_image_path(vals[0], "")
                    
                    # Close window and reload tree
                    img_win.destroy()
                    load_to_tree()
                    
                    info_var.set("Görsel silindi.")
                    messagebox.showinfo("Başarılı", "Görsel başarıyla silindi!")
                    
                except Exception as e:
                    messagebox.showerror("Hata", f"Görsel silinirken hata: {str(e)}")
        
        CTkButton(img_win, text="Görseli Sil", command=delete_image, 
                 width=140, height=35, fg_color="#E53935", 
                 hover_color="#C62828").pack(pady=10)
        
    except Exception as e:
        error_msg = f"Görsel yüklenemedi: {str(e)}"
        info_var.set(error_msg)
        messagebox.showerror("Hata", error_msg)
        
# def generate_all_barcodes():
#     """Generate barcodes for all items"""
#     try:
#         rows = read_all_rows()
#         generated_count = 0
        
#         for row in rows:
#             item_code = row[1]
#             item_name = row[11] if len(row) > 11 else row[3]
            
#             barcode_path = generate_barcode(item_code, item_name)
#             if barcode_path:
#                 generated_count += 1
        
#         info_var.set(f"Toplam {generated_count} barkod oluşturuldu.")
#         messagebox.showinfo("Başarılı", f"Toplam {generated_count} barkod oluşturuldu.\nBarkodlar 'barcodes' klasöründe.")
        
#     except Exception as e:
#         messagebox.showerror("Hata", f"Barkod oluşturma hatası: {str(e)}")

def generate_all_qrcodes():
    """Generate QR codes for all items with complete information"""
    try:
        rows = read_all_rows()
        generated_count = 0
        
        for row in rows:
            item_code = row[1]
            item_data = {
                'id': row[0],
                'internal_code': row[2],
                'inventory_item': row[3],
                'turkish_name': row[11] if len(row) > 11 else row[3],
                'campus': row[4],
                'floor': row[5],
                'room': row[6],
                'category': row[7],
                'status': row[8],
                'date': row[9],
                'color': row[10] if len(row) > 10 else ""
            }
            
            qr_path = generate_qrcode(item_data, item_code)
            if qr_path:
                generated_count += 1
        
        info_var.set(f"Toplam {generated_count} QR kod oluşturuldu.")
        messagebox.showinfo("Başarılı", f"Toplam {generated_count} QR kod oluşturuldu.\nTüm bilgiler QR kodlar içinde mevcut.")
        
    except Exception as e:
        messagebox.showerror("Hata", f"QR kod oluşturma hatası: {str(e)}")

# ---------- Room QR Code Functions ----------
def generate_room_qrcode(room_name, items_in_room):
    """Generate QR code for a room containing all items in that room (optimized for large lists)"""
    try:
        # Use ERROR_CORRECT_L for maximum data capacity within standard QR limits
        qr = qrcode.QRCode(
            version=None,
            error_correction=qrcode.constants.ERROR_CORRECT_L,
            box_size=8,
            border=4,
        )
        
        # Room Summary
        qr_data = f"AVRASYA UNIVERSITESI - ODA ENVANTERI\n"
        qr_data += f"Oda: {room_name}\n"
        qr_data += f"Toplam Demirbaş: {len(items_in_room)} Adet\n"
        qr_data += "-" * 20 + "\n"
        
        # Categorized Summary
        category_counts = {}
        for item in items_in_room:
            cat = item.get('category', 'Diğer')
            category_counts[cat] = category_counts.get(cat, 0) + 1
        
        for cat, count in sorted(category_counts.items()):
            qr_data += f"{cat}: {count}\n"
        
        qr_data += "-" * 20 + "\n"
        # Group items by name to save space (User's excellent suggestion)
        item_grouping = {}
        for item in items_in_room:
            name = item['turkish_name']
            item_grouping[name] = item_grouping.get(name, 0) + 1
        
        qr_data += "DEMİRBAŞ LİSTESİ (Özet):\n"
        
        # Display summarized list: "Name - X Adet"
        for idx, (name, count) in enumerate(sorted(item_grouping.items()), 1):
            qr_data += f"{idx}. {name}: {count} Adet\n"
        
        qr_data += "-" * 20 + "\n"
        qr_data += f"Tarih: {strftime('%Y-%m-%d')}"
        
        qr.add_data(qr_data)
        qr.make(fit=True)
        
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # Safe filename
        safe_room_name = "".join(c for c in room_name if c.isalnum() or c in (' ', '_', '-')).strip()
        qr_path = os.path.join(QRCODE_DIR, f"room_{safe_room_name}.png")
        qr_img.save(qr_path)
        
        return qr_path
    except Exception as e:
        print(f"Room QR code generation error: {e}")
        return None

def show_room_qrcode():
    """Generate and show QR code for a specific room"""
    # Get list of all rooms
    rows = read_all_rows()
    rooms = sorted(set(r[6] for r in rows if r[6]))  # Column 6 is Room/Oda
    
    if not rooms:
        messagebox.showinfo("Bilgi", "Henüz hiç oda kaydı yok.")
        return
    
    # Create selection window
    select_win = CTkToplevel(w)
    select_win.title("Oda Seçin")
    select_win.geometry("400x1000")
    
    CTkLabel(select_win, text="ODA/SINIF SEÇİN", 
             font=("Arial", 16, "bold")).pack(pady=10)
    
    CTkLabel(select_win, text="QR kod oluşturmak için bir oda seçin:", 
             font=("Arial", 12)).pack(pady=5)
    
    # Create scrollable frame for rooms
    scroll_frame = CTkScrollableFrame(select_win, width=350, height=350)
    scroll_frame.pack(pady=10, padx=20, fill="both", expand=True)
    
    selected_room = StringVar()
    
    for room in rooms:
        # Count items in this room
        room_items = [r for r in rows if r[6] == room]
        item_count = len(room_items)
        
        CTkRadioButton(
            scroll_frame,
            text=f"{room} ({item_count} demirbaş)",
            variable=selected_room,
            value=room,
            font=("Arial", 11)
        ).pack(anchor="w", pady=3, padx=10)
    
    def generate_selected():
        room = selected_room.get()
        if not room:
            messagebox.showwarning("Uyarı", "Lütfen bir oda seçin!")
            return
        
        # Get all items in this room
        items_in_room = []
        for row in rows:
            if row[6] == room:
                items_in_room.append({
                    'code': row[1],
                    'turkish_name': row[11] if len(row) > 11 else row[3],
                    'category': row[7],
                    'status': row[8],
                    'color': row[10] if len(row) > 10 else ""
                })
        
        # Generate QR code
        qr_path = generate_room_qrcode(room, items_in_room)
        
        if qr_path:
            select_win.destroy()
            show_room_qr_window(room, items_in_room, qr_path)
        else:
            messagebox.showerror("Hata", "QR kod oluşturulamadı!")
    
    CTkButton(
        select_win,
        text="QR Kod Oluştur",
        command=generate_selected,
        width=200,
        height=60,
        fg_color="#2962FF",
        hover_color="#0039CB",
        font=("Arial", 14, "bold")
    ).pack(pady=10)

def show_room_qr_window(room_name, items, qr_path):
    """Display the room QR code in a new window"""
    qr_win = CTkToplevel(w)
    qr_win.title(f"Oda QR Kodu - {room_name}")
    qr_win.geometry("700x800")
    
    # Title
    CTkLabel(qr_win, text="ODA/SINIF QR KODU", 
             font=("Arial", 20, "bold")).pack(pady=10)
    
    # Room info
    info_frame = CTkFrame(qr_win)
    info_frame.pack(pady=10, padx=20, fill="x")
    
    CTkLabel(info_frame, text=f"📍 Oda: {room_name}", 
             font=("Arial", 14, "bold")).pack(anchor="w", pady=3)
    CTkLabel(info_frame, text=f"📦 Toplam Demirbaş: {len(items)} adet", 
             font=("Arial", 12)).pack(anchor="w", pady=2)
    
    # Category breakdown
    category_counts = {}
    for item in items:
        cat = item.get('category', 'Diğer')
        category_counts[cat] = category_counts.get(cat, 0) + 1
    
    cat_text = "Kategoriler: "
    cat_text += ", ".join([f"{cat} ({count})" for cat, count in sorted(category_counts.items())])
    CTkLabel(info_frame, text=cat_text, 
             font=("Arial", 10), text_color="gray").pack(anchor="w", pady=2)
    
    # QR Code
    try:
        qr_img = Image.open(qr_path)
        qr_img = qr_img.resize((350, 350), Image.LANCZOS)
        qr_ctk_image = CTkImage(light_image=qr_img, dark_image=qr_img, size=(350, 350))
        
        img_label = CTkLabel(qr_win, image=qr_ctk_image, text="")
        img_label.image = qr_ctk_image
        img_label.pack(pady=15)
    except Exception as e:
        CTkLabel(qr_win, text=f"QR kod yüklenemedi: {str(e)}", 
                 font=("Arial", 12)).pack(pady=20)
    
    # Instructions
    inst_frame = CTkFrame(qr_win)
    inst_frame.pack(pady=10, padx=20, fill="x")
    
    CTkLabel(inst_frame, text="📱 KULLANIM TALİMATI:", 
             font=("Arial", 12, "bold")).pack(anchor="w", pady=3)
    CTkLabel(inst_frame, text="1. Bu QR kodu yazdırın", 
             font=("Arial", 10)).pack(anchor="w", padx=20, pady=1)
    CTkLabel(inst_frame, text="2. Oda/sınıf kapısına yapıştırın", 
             font=("Arial", 10)).pack(anchor="w", padx=20, pady=1)
    CTkLabel(inst_frame, text="3. Mobil ile QR okuyucu kullanarak tarayın", 
             font=("Arial", 10)).pack(anchor="w", padx=20, pady=1)
    CTkLabel(inst_frame, text="4. Odadaki tüm demirbaşları görün", 
             font=("Arial", 10)).pack(anchor="w", padx=20, pady=1)
    
    # Print button
    CTkButton(
        qr_win,
        text="🖨️ QR Kodu Yazdır",
        command=lambda: print_room_qrcode(qr_path, room_name, items),
        width=200,
        height=40,
        fg_color="#2962FF",
        hover_color="#0039CB",
        font=("Arial", 14, "bold")
    ).pack(pady=10)

def print_room_qrcode(qr_path, room_name, items):
    """Print room QR code with information"""
    try:
        pdf = FPDF()
        pdf.add_page()
        
        # Add university logo if exists
        if os.path.exists(LOGO_FILE):
            pdf.image(LOGO_FILE, x=85, y=10, w=40)
            pdf.set_y(35)
        else:
            pdf.set_y(10)
        
        # Title
        pdf.set_font("Arial", 'B', 18)
        pdf.set_text_color(0, 51, 102)
        pdf.cell(200, 10, "AVRASYA UNIVERSITESI", ln=True, align='C')
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(200, 8, "ODA/SINIF DEMIRBAS LISTESI", ln=True, align='C')
        pdf.ln(5)
        
        # Room name
        pdf.set_font("Arial", 'B', 14)
        pdf.set_text_color(0, 0, 0)
        pdf.cell(200, 8, f"ODA: {room_name}", ln=True, align='C')
        pdf.set_font("Arial", '', 12)
        pdf.cell(200, 7, f"Toplam Demirbas: {len(items)} Adet", ln=True, align='C')
        pdf.ln(5)
        
        # Separator
        pdf.set_draw_color(0, 51, 102)
        pdf.line(20, pdf.get_y(), 190, pdf.get_y())
        pdf.ln(10)
        
        # QR Code (centered)
        pdf.image(qr_path, x=55, y=pdf.get_y(), w=100)
        pdf.ln(110)
        
        # Instructions
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(200, 7, "KULLANIM TALİMATI:", ln=True)
        pdf.set_font("Arial", '', 10)
        pdf.multi_cell(0, 6, 
            "1. Bu sayfayı renkli yazdırın\n"
            "2. Oda/sınıf kapısına görünür bir yere yapıştırın\n"
            "3. QR kodu mobil cihazınızla tarayın\n"
            "4. Odadaki tüm demirbaş listesini görüntüleyin")
        pdf.ln(5)
        
        # Item list
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(200, 7, f"DEMIRBAS LISTESI ({len(items)} Adet):", ln=True)
        pdf.set_font("Arial", '', 9)
        
        for idx, item in enumerate(items, 1):
            item_text = f"{idx}. {item['turkish_name']} - {item['category']} ({item['status']})"
            pdf.cell(200, 5, item_text, ln=True)
        
        # Footer
        pdf.ln(5)
        pdf.set_font("Arial", 'I', 8)
        pdf.set_text_color(100, 100, 100)
        pdf.cell(200, 5, f"Olusturma: {strftime('%Y-%m-%d %H:%M')}", ln=True, align='C')
        pdf.cell(200, 5, "Demirbas Yonetimi - Avrasya Universitesi", ln=True, align='C')
        
        # Save PDF
        filename = f"Oda_QR_{room_name.replace(' ', '_')}.pdf"
        pdf.output(filename)
        
        info_var.set(f"Oda QR belgesi oluşturuldu: {filename}")
        messagebox.showinfo("Başarılı", f"Oda QR belgesi oluşturuldu:\n{filename}\n\nBu dosyayı yazdırıp oda kapısına yapıştırabilirsiniz!")
        
    except Exception as e:
        messagebox.showerror("Hata", f"Yazdırma hatası: {str(e)}")

def generate_all_room_qrcodes():
    """Generate QR codes for all rooms"""
    try:
        rows = read_all_rows()
        rooms = set(r[6] for r in rows if r[6])
        
        if not rooms:
            messagebox.showinfo("Bilgi", "Henüz hiç oda kaydı yok.")
            return
        
        generated_count = 0
        
        for room in rooms:
            # Get all items in this room
            items_in_room = []
            for row in rows:
                if row[6] == room:
                    items_in_room.append({
                        'code': row[1],
                        'turkish_name': row[11] if len(row) > 11 else row[3],
                        'category': row[7],
                        'status': row[8],
                        'color': row[10] if len(row) > 10 else ""
                    })
            
            qr_path = generate_room_qrcode(room, items_in_room)
            if qr_path:
                generated_count += 1
        
        info_var.set(f"Toplam {generated_count} oda QR kodu oluşturuldu.")
        messagebox.showinfo("Başarılı", 
                           f"Toplam {generated_count} oda için QR kod oluşturuldu!\n\n"
                           f"QR kodlar 'qrcodes' klasöründe.\n"
                           f"Her QR kod o odadaki tüm demirbaşları içerir.")
        
    except Exception as e:
        messagebox.showerror("Hata", f"Oda QR kodları oluşturma hatası: {str(e)}")

def show_grouped_item_qrcodes():
    """Generate QR codes for groups of identical items in a room"""
    rows = read_all_rows()
    rooms = sorted(set(r[6] for r in rows if r[6]))
    
    if not rooms:
        messagebox.showinfo("Bilgi", "Henüz hiç oda kaydı yok.")
        return
    
    # Selection window for room
    sel_win = CTkToplevel(w)
    sel_win.title("Oda Seç - Grup QR")
    sel_win.geometry("400x600")
    
    CTkLabel(sel_win, text="GRUP QR OLUŞTUR", font=("Arial", 16, "bold")).pack(pady=10)
    CTkLabel(sel_win, text="Oda seçerek benzer eşyalar için QR oluşturun:").pack(pady=5)
    
    scroll = CTkScrollableFrame(sel_win, width=350, height=400)
    scroll.pack(pady=10, padx=20, fill="both", expand=True)
    
    for room in rooms:
        def make_handler(r=room):
            return lambda: generate_groups_for_room(r, sel_win)
            
        CTkButton(scroll, text=room, command=make_handler(room)).pack(pady=2, fill="x", padx=10)

def generate_groups_for_room(room_name, parent_win):
    """Group items in a room and generate QR codes for each group"""
    rows = read_all_rows()
    room_items = [r for r in rows if r[6] == room_name]
    
    if not room_items:
        return
        
    # Group by name
    groups = {}
    for r in room_items:
        name = r[11] if len(r) > 11 else r[3]
        if name not in groups:
            groups[name] = []
        groups[name].append(r)
        
    # Show groups window
    group_win = CTkToplevel(parent_win)
    group_win.title(f"Gruplar - {room_name}")
    group_win.geometry("500x600")
    
    CTkLabel(group_win, text=f"📍 {room_name} - Eşya Grupları", font=("Arial", 14, "bold")).pack(pady=10)
    
    list_frame = CTkScrollableFrame(group_win, width=450, height=450)
    list_frame.pack(pady=10, padx=20, fill="both", expand=True)
    
    for name, items in groups.items():
        frame = CTkFrame(list_frame)
        frame.pack(pady=5, fill="x", padx=5)
        
        CTkLabel(frame, text=f"{name} ({len(items)} Adet)", font=("Arial", 12, "bold")).pack(side="left", padx=10)
        
        def make_qr_handler(n=name, its=items):
            return lambda: create_and_show_group_qr(room_name, n, its)
            
        CTkButton(frame, text="QR Oluştur", width=100, command=make_qr_handler(name, items),
                 fg_color="#2962FF", hover_color="#0039CB", font=("Arial", 11, "bold")).pack(side="right", padx=10, pady=5)

def create_and_show_group_qr(room_name, group_name, items):
    """Create a single QR code for a group of items and display it"""
    try:
        # Create QR data
        qr_data = f"AVRASYA UNIVERSITESI - GRUP BILGISI\n"
        qr_data += f"Oda: {room_name}\n"
        qr_data += f"Eşya: {group_name}\n"
        qr_data += f"Toplam Adet: {len(items)}\n"
        qr_data += "-" * 20 + "\n"
        qr_data += "DEMİRBAŞ NO LİSTESİ:\n"
        
        for idx, it in enumerate(items, 1):
            qr_data += f"{idx}. {it[1]}\n" # Column 1 is Item Code
            
        qr_data += "-" * 20 + "\n"
        qr_data += f"Oluşturma: {strftime('%Y-%m-%d')}"
        
        # Generate QR
        qr = qrcode.QRCode(version=None, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
        qr.add_data(qr_data)
        qr.make(fit=True)
        qr_img = qr.make_image(fill_color="black", back_color="white")
        
        # Save file
        safe_name = "".join(c for c in f"{room_name}_{group_name}" if c.isalnum() or c in (' ', '_')).strip()
        path = os.path.join(QRCODE_DIR, f"group_{safe_name}.png")
        qr_img.save(path)
        
        # Display window
        qr_win = CTkToplevel(w)
        qr_win.title(f"Grup QR - {group_name}")
        qr_win.geometry("500x700")
        
        CTkLabel(qr_win, text=f"GRUP QR KODU", font=("Arial", 18, "bold")).pack(pady=10)
        CTkLabel(qr_win, text=f"{room_name} - {group_name}", font=("Arial", 14)).pack(pady=2)
        
        img = Image.open(path)
        img = img.resize((350, 350), Image.LANCZOS)
        ctk_img = CTkImage(light_image=img, dark_image=img, size=(350, 350))
        
        label = CTkLabel(qr_win, image=ctk_img, text="")
        label.image = ctk_img
        label.pack(pady=20)
        
        CTkLabel(qr_win, text=f"Bu QR کد {len(items)} قلم کالای مشابه را نمایندگی می‌کند.", font=("Arial", 10)).pack(pady=5)
        
        def print_group_pdf():
            pdf = FPDF()
            pdf.add_page()
            
            if os.path.exists(LOGO_FILE):
                pdf.image(LOGO_FILE, x=85, y=10, w=40)
                pdf.set_y(35)
            else:
                pdf.set_y(10)
                
            pdf.set_font("Arial", 'B', 18)
            pdf.cell(200, 15, "AVRASYA UNIVERSITESI", ln=True, align='C')
            pdf.set_font("Arial", 'B', 16)
            pdf.cell(200, 10, "GRUP DEMIRBAS QR KODU", ln=True, align='C')
            pdf.ln(10)
            
            pdf.set_font("Arial", '', 14)
            pdf.cell(200, 8, f"Oda: {room_name}", ln=True)
            pdf.cell(200, 8, f"Eşya: {group_name}", ln=True)
            pdf.cell(200, 8, f"Toplam Adet: {len(items)}", ln=True)
            pdf.ln(10)
            
            pdf.image(path, x=55, y=pdf.get_y(), w=100)
            pdf.ln(110)
            
            pdf.set_font("Arial", 'B', 12)
            pdf.cell(200, 10, "KOD LISTESI / DEMIRBAS NO:", ln=True)
            pdf.set_font("Arial", '', 10)
            for it in items:
                pdf.cell(200, 6, f"- {it[1]}", ln=True)
                
            pdf_out = f"Grup_QR_{safe_name}.pdf"
            pdf.output(pdf_out)
            messagebox.showinfo("Başarılı", f"Grup QR PDF oluşturuldu:\n{pdf_out}")
            
        CTkButton(qr_win, text="🖨️ PDF Oluştur / Yazdır", command=print_group_pdf, 
                 width=200, height=40, fg_color="#2962FF", font=("Arial", 14, "bold")).pack(pady=10)
        
    except Exception as e:
        messagebox.showerror("Hata", f"Grup QR hatası: {str(e)}")

# def generate_all_detailed_barcodes():
#     """Generate detailed barcodes for all items"""
#     try:
#         rows = read_all_rows()
#         generated_count = 0
        
#         for row in rows:
#             item_code = row[1]
#             item_data = {
#                 'turkish_name': row[11] if len(row) > 11 else row[3],
#                 'campus': row[4],
#                 'room': row[6]
#             }
            
#             barcode_path = generate_detailed_barcode(item_data, item_code)
#             if barcode_path:
#                 generated_count += 1
        
#         info_var.set(f"Toplam {generated_count} detaylı barkod oluşturuldu.")
#         messagebox.showinfo("Başarılı", f"Toplam {generated_count} detaylı barkod oluşturuldu.")
        
#     except Exception as e:
#         messagebox.showerror("Hata", f"Detaylı barkod oluşturma hatası: {str(e)}")

# ---------- Import from Excel Function ----------
def import_from_excel():
    file_path = filedialog.askopenfilename(
        title="Excel dosyası seçin",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    
    if not file_path:
        return
    
    try:
        excel_file = pd.ExcelFile(file_path)
        imported_count = 0
        
        for sheet_name in excel_file.sheet_names:
            # Read without header first to detect structure
            df_raw = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            print(f"Processing sheet: {sheet_name}")
            print(f"Raw data shape: {df_raw.shape}")
            
            # Find header row by looking for Turkish column names
            header_row = 0
            for i in range(min(5, len(df_raw))):  # Check first 5 rows
                row_values = [str(cell).lower().strip() for cell in df_raw.iloc[i] if pd.notna(cell)]
                print(f"Row {i} values: {row_values}")
                
                if any('adı' in val or 'cinsi' in val or 'renk' in val or 'adet' in val for val in row_values):
                    header_row = i
                    print(f"Found header at row: {i}")
                    break
            
            # Read with correct header
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
            df.columns = [str(col).strip() for col in df.columns]
            
            print(f"Columns after header detection: {list(df.columns)}")
            
            # Map columns - IMPROVED DETECTION
            column_mapping = {}
            for col in df.columns:
                col_lower = str(col).lower().strip()
                print(f"Mapping column: '{col}' -> '{col_lower}'")
                
                if any(keyword in col_lower for keyword in ['adı', 'adi', 'ad', 'isim', 'malzeme adı']):
                    column_mapping['turkish_name'] = col
                    print(f"✓ Turkish Name: {col}")
                elif any(keyword in col_lower for keyword in ['cinsi', 'cins', 'kategori', 'tür', 'tip']):
                    column_mapping['category'] = col
                    print(f"✓ Category: {col}")
                elif 'renk' in col_lower:
                    column_mapping['color'] = col
                    print(f"✓ Color: {col}")
                elif any(keyword in col_lower for keyword in ['adet', 'miktar', 'sayı', 'sayi', 'quantity']):
                    column_mapping['quantity'] = col
                    print(f"✓ Quantity: {col}")
            
            # Process each row
            for index, row in df.iterrows():
                # Skip header row
                if index == header_row:
                    continue
                
                # Extract data using mapping
                turkish_name = ""
                if 'turkish_name' in column_mapping:
                    turkish_name = str(row[column_mapping['turkish_name']]).strip() if pd.notna(row[column_mapping['turkish_name']]) else ""
                
                # If no Turkish name found, skip the row
                if not turkish_name:
                    print(f"Skipping row {index} - no Turkish name found")
                    continue
                
                print(f"Processing item: '{turkish_name}'")
                
                # Extract other fields
                category = ""
                if 'category' in column_mapping:
                    category = str(row[column_mapping['category']]).strip() if pd.notna(row[column_mapping['category']]) else ""
                
                color = ""
                if 'color' in column_mapping:
                    color = str(row[column_mapping['color']]).strip() if pd.notna(row[column_mapping['color']]) else ""
                
                quantity = 1
                if 'quantity' in column_mapping:
                    try:
                        qty_val = row[column_mapping['quantity']]
                        if pd.notna(qty_val):
                            quantity = int(float(qty_val))
                    except:
                        quantity = 1
                
                # Create records based on quantity
                for i in range(quantity):
                    # Generate safe name for code
                    safe_name = "".join(c for c in turkish_name[:8] if c.isalnum()).upper()
                    if not safe_name:
                        safe_name = "ITEM"
                    
                    item_code = f"{sheet_name[:3].upper()}_{safe_name}_{imported_count + 1}"
                    internal_code = f"INT_{sheet_name[:2].upper()}_{imported_count + 1}"
                    
                    data = [
                        str(uuid4()),
                        item_code,
                        internal_code,
                        turkish_name,  # Inventory Item
                        "Pelitli",
                        "Kat 0",
                        sheet_name,
                        category if category else "Diğer",
                        "Kullanilmiş",
                        strftime('%Y-%m-%d'),
                        color,
                        turkish_name,  # Turkish Name (same as inventory item)
                        ""  # Image_Path - empty for imported items
                    ]
                    
                    print(f"Adding record: '{turkish_name}' | Color: '{color}' | Category: '{category}' | Quantity: {quantity}")
                    
                    append_row(data)
                    imported_count += 1
        
        load_to_tree()
        refresh_dynamic_choices()
        
        info_var.set(f"Başarıyla {imported_count} öğe {len(excel_file.sheet_names)} sayfadan içe aktarıldı")
        messagebox.showinfo("İçe Aktarma Başarılı", 
                           f"{imported_count} öğe {len(excel_file.sheet_names)} sayfadan içe aktarıldı!\n\n"
                           f"Dosya: {os.path.basename(file_path)}")
        
    except Exception as e:
        error_msg = f"Dosya içe aktarılırken hata: {str(e)}"
        print(f"Error: {error_msg}")
        info_var.set(error_msg)
        messagebox.showerror("İçe Aktarma Hatası", error_msg)

# ---------- Simple Import Function ----------

def simple_import_excel():
    file_path = filedialog.askopenfilename(
        title="Excel dosyası seçin",
        filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
    )
    
    if not file_path:
        return
    
    try:
        # Read the Excel file
        excel_file = pd.ExcelFile(file_path)
        imported_count = 0
        
        for sheet_name in excel_file.sheet_names:
            # Read without header assumption
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
            
            print(f"Processing sheet: {sheet_name}")
            print(f"Data shape: {df.shape}")
            
            # Try to find header row
            header_row = 0
            for i in range(min(3, len(df))):  # Check first 3 rows
                row_values = [str(cell).lower() for cell in df.iloc[i] if pd.notna(cell)]
                if any('adı' in val or 'cinsi' in val or 'renk' in val for val in row_values):
                    header_row = i
                    print(f"Found header at row: {i}")
                    break
            
            # Read again with correct header
            df = pd.read_excel(file_path, sheet_name=sheet_name, header=header_row)
            df.columns = [str(col).strip() for col in df.columns]
            
            print(f"Columns after header detection: {list(df.columns)}")
            
            # Process each row
            for index, row in df.iterrows():
                # Skip if this is the header row
                if index == header_row:
                    continue
                
                # Look for data in common column positions
                turkish_name = ""
                category = ""
                color = ""
                quantity = 1
                
                # Try different column positions for Turkish name
                for col in df.columns:
                    col_lower = str(col).lower()
                    if pd.notna(row[col]) and str(row[col]).strip():
                        cell_value = str(row[col]).strip()
                        
                        if not turkish_name and ('adı' in col_lower or 'adi' in col_lower or col_lower in ['ad', 'isim', 'malzeme adı']):
                            turkish_name = cell_value
                        elif not category and ('cinsi' in col_lower or 'kategori' in col_lower or col_lower in ['cins', 'tür']):
                            category = cell_value
                        elif not color and ('renk' in col_lower):
                            color = cell_value
                        elif quantity == 1 and ('adet' in col_lower or 'miktar' in col_lower):
                            try:
                                quantity = int(cell_value)
                            except:
                                quantity = 1
                
                # If still no Turkish name, use first non-empty column
                if not turkish_name:
                    for col in df.columns:
                        if pd.notna(row[col]) and str(row[col]).strip():
                            turkish_name = str(row[col]).strip()
                            break
                
                if not turkish_name:
                    continue
                
                print(f"Importing: {turkish_name}")
                
                # Create records
                for i in range(quantity):
                    safe_name = "".join(c for c in turkish_name[:8] if c.isalnum()).upper()
                    item_code = f"{sheet_name[:3].upper()}_{safe_name}_{imported_count + 1}"
                    internal_code = f"INT_{sheet_name[:2].upper()}_{imported_count + 1}"
                    
                    data = [
                        str(uuid4()),
                        item_code,
                        internal_code,
                        turkish_name,
                        "Pelitli",
                        "Kat 0",
                        sheet_name,
                        category if category else "Diğer",
                        "Kullanilmiş",
                        strftime('%Y-%m-%d'),
                        color,
                        turkish_name,
                        ""  # Image_Path - empty for imported items
                    ]
                    
                    append_row(data)
                    imported_count += 1
        
        load_to_tree()
        refresh_dynamic_choices()
        
        info_var.set(f"Başarıyla {imported_count} öğe içe aktarıldı")
        messagebox.showinfo("Başarılı", f"{imported_count} öğe içe aktarıldı!")
        
    except Exception as e:
        error_msg = f"Hata: {str(e)}"
        print(f"Error: {error_msg}")
        messagebox.showerror("Hata", error_msg)

# ---------- Other Functions (remain the same) ----------
def now_time():
    return strftime('%H:%M:%S')

def clean_text(s: str) -> str:
    return re.sub(r"[^\x00-\xFF]+", "", s or "")

def read_all_rows():
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(list(row))
    return rows

def append_row(data):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    ws.append(data)
    wb.save(EXCEL_FILE)

def update_row_by_id(row_id, new_vals_without_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(row_id):
            for i, val in enumerate(new_vals_without_id, start=1):
                row[i].value = val
            break
    wb.save(EXCEL_FILE)

def delete_row_by_id(row_id):
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    for row in ws.iter_rows(min_row=2):
        if str(row[0].value) == str(row_id):
            ws.delete_rows(row[0].row)
            break
    wb.save(EXCEL_FILE)

# ---------- UI Data ----------
current_view_rows = []
edit_mode = False
edited_tree_iid = None
edited_row_id = None

# ---------- Update Comboboxes from Data ----------
def refresh_dynamic_choices():
    rows = read_all_rows()
    campuses = sorted({r[4] for r in rows if r[4]})
    floors = sorted({r[5] for r in rows if r[5]})
    rooms = sorted({r[6] for r in rows if r[6]})
    categories = sorted({r[7] for r in rows if r[7]})
    statuses = sorted({r[8] for r in rows if r[8]})
    colors = sorted({r[10] for r in rows if r[10]})

    campus_cb.configure(values=tuple(sorted(set(list(campus_cb.cget('values')) + campuses))))
    floor_cb.configure(values=tuple(sorted(set(list(floor_cb.cget('values')) + floors))))
    room_cb.configure(values=tuple(sorted(set(list(room_cb.cget('values')) + rooms))))
    category_cb.configure(values=tuple(sorted(set(list(category_cb.cget('values')) + categories))))
    status_cb.configure(values=tuple(sorted(set(list(status_cb.cget('values')) + statuses))))
    color_cb.configure(values=tuple(sorted(set(list(color_cb.cget('values')) + colors))))

# ---------- Load to Table ----------
def load_to_tree(rows=None):
    global current_view_rows
    if rows is None:
        rows = read_all_rows()
    current_view_rows = rows[:]
    tree.delete(*tree.get_children())
    for r in rows:
        tree.insert("", "end", values=r)
    result_count_var.set(f"Görüntüleme sayısı: {len(rows)}")

# ---------- Save ----------
def save_record():
    data = [
        str(uuid4()),
        code_var.get().strip(),
        internal_code_var.get().strip(),
        inventory_var.get().strip(),
        campus_var.get().strip(),
        floor_var.get().strip(),
        room_var.get().strip(),
        category_var.get().strip(),
        status_var.get().strip(),
        date_var.get().strip(),
        color_var.get().strip(),
        turkish_name_var.get().strip(),
        ""  # Image_Path - empty by default
    ]
    if not data[3] or not data[4]:
        info_var.set("Demirbaş Öğesi ve Kampüs zorunludur.")
        return
    append_row(data)
    load_to_tree()
    refresh_dynamic_choices()
    clear_fields()
    info_var.set("Kaydedildi ve alanlar temizlendi.")

# ---------- Select for Editing ----------
def edit_selected():
    global edit_mode, edited_tree_iid, edited_row_id
    sel = tree.selection()
    if not sel:
        info_var.set("Hiçbir satır seçilmedi.")
        return
    edited_tree_iid = sel[0]
    vals = tree.item(sel[0])['values']
    edited_row_id = vals[0]

    code_var.set(vals[1]); internal_code_var.set(vals[2]); inventory_var.set(vals[3])
    campus_var.set(vals[4]); floor_var.set(vals[5]); room_var.set(vals[6])
    category_var.set(vals[7]); status_var.set(vals[8]); date_var.set(vals[9])
    color_var.set(vals[10] if len(vals) > 10 else "")
    turkish_name_var.set(vals[11] if len(vals) > 11 else "")

    edit_mode = True
    info_var.set("Düzenleme modundasınız. Değişiklikleri Kaydet Düzenlendi ile kaydedin.")

def save_edited():
    global edit_mode, edited_row_id, edited_tree_iid
    if not edit_mode or not edited_row_id:
        return
    
    # Get existing image path from the selected row
    existing_image_path = ""
    if edited_tree_iid:
        vals = tree.item(edited_tree_iid)['values']
        existing_image_path = vals[12] if len(vals) > 12 else ""
    
    new_vals = [
        code_var.get().strip(),
        internal_code_var.get().strip(),
        inventory_var.get().strip(),
        campus_var.get().strip(),
        floor_var.get().strip(),
        room_var.get().strip(),
        category_var.get().strip(),
        status_var.get().strip(),
        date_var.get().strip(),
        color_var.get().strip(),
        turkish_name_var.get().strip(),
        existing_image_path  # Preserve existing image path
    ]
    update_row_by_id(edited_row_id, new_vals)
    edit_mode = False
    edited_row_id = None
    load_to_tree()
    refresh_dynamic_choices()
    clear_fields()
    info_var.set("Düzenleme kaydedildi ve alanlar temizlendi.")

# ---------- Delete ----------
def delete_selected():
    sel = tree.selection()
    if not sel:
        info_var.set("Hiçbir satır seçilmedi.")
        return
    vals = tree.item(sel[0])['values']
    row_id = vals[0]
    delete_row_by_id(row_id)
    load_to_tree()
    info_var.set("Silindi.")

# ---------- Filter ----------
def filter_view():
    y = filter_campus_var.get().strip().lower()
    o = filter_room_var.get().strip().lower()
    d = filter_inventory_var.get().strip().lower()
    k = filter_code_var.get().strip().lower()
    ik = filter_internal_code_var.get().strip().lower()
    tn = filter_turkish_name_var.get().strip().lower()
    col = filter_color_var.get().strip().lower()
    cat = filter_category_var.get().strip().lower()

    rows = read_all_rows()
    out = []
    for r in rows:
        ok = True
        if y and y not in str(r[4]).lower(): ok = False
        if o and o not in str(r[6]).lower(): ok = False
        if d and d not in str(r[3]).lower(): ok = False
        if k and k not in str(r[1]).lower(): ok = False
        if ik and ik not in str(r[2]).lower(): ok = False
        if cat and cat not in str(r[7]).lower(): ok = False
        if tn and tn not in str(r[11] if len(r) > 11 else "").lower(): ok = False
        if col and col not in str(r[10] if len(r) > 10 else "").lower(): ok = False
        if ok: out.append(r)
    load_to_tree(out)

    room_total = len(out) if o else "-"
    result_count_var.set(f"Display count: {len(out)} | Room total: {room_total}")

def show_all():
    filter_campus_var.set(""); filter_room_var.set(""); filter_inventory_var.set("")
    filter_turkish_name_var.set(""); filter_color_var.set(""); filter_category_var.set("")
    load_to_tree()

# ---------- Export from Current View ----------
def export_view_to_csv():
    with open("export_view.csv", "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)
        for r in current_view_rows:
            w.writerow(r)
    info_var.set("CSV oluşturuldu: export_view.csv")

def export_view_to_pdf():
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=9)
    pdf.cell(200, 8, txt="Avrasya University - Inventory (View)", ln=True)
    pdf.ln(2)
    for r in current_view_rows:
        line = " | ".join(str(x) for x in r)
        pdf.cell(200, 6, txt=clean_text(line), ln=True)
    pdf.output("export_view.pdf")
    info_var.set("PDF oluşturuldu: export_view.pdf")

# ---------- Statistical Summary ----------
def make_summary(rows=None):
    if rows is None:
        rows = current_view_rows if current_view_rows else read_all_rows()

    N = len(rows)
    by_y = Counter(r[4] for r in rows if r[4])
    by_k = Counter(r[7] for r in rows if r[7])
    by_s = Counter(r[8] for r in rows if r[8])
    by_o = Counter(r[6] for r in rows if r[6])
    by_color = Counter(r[10] for r in rows if len(r) > 10 and r[10])

    return {"total": N, "by_campus": by_y, "by_category": by_k, "by_status": by_s, "by_room": by_o, "by_color": by_color}

def show_summary():
    s = make_summary()
    win = CTkToplevel(w); win.title("Demirbaş Özeti"); win.geometry("500x600")

    txt = CTkTextbox(win, font=("Courier", 10))
    def section(title, counter):
        txt.insert(END, f"{title}\n")
        for k, v in counter.most_common():
            txt.insert(END, f"  - {k}: {v}\n")
        txt.insert(END, "\n")

    txt.insert(END, f"Total (All current displays): {s['total']}\n\n")
    section("Campus:", s["by_campus"])
    section("Category:", s["by_category"])
    section("Status:", s["by_status"])
    section("Room:", s["by_room"])
    section("Color:", s["by_color"])
    txt.pack(expand=True, fill="both")

def save_summary_files():
    s = make_summary()

    # PDF
    pdf = FPDF(); pdf.add_page(); pdf.set_font("Arial", size=11)
    pdf.cell(200, 8, txt="Avrasya University - Demirbaş Özeti", ln=True)
    pdf.ln(2)
    pdf.cell(200, 8, txt=f"Total: {s['total']}", ln=True); pdf.ln(2)

    def pdf_block(title, counter):
        pdf.cell(200, 8, txt=clean_text(title), ln=True)
        for k, v in counter.most_common():
            pdf.cell(200, 7, txt=clean_text(f"  - {k}: {v}"), ln=True)
        pdf.ln(2)

    pdf_block("Campus:", s["by_campus"])
    pdf_block("Category:", s["by_category"])
    pdf_block("Status:", s["by_status"])
    pdf_block("Room:", s["by_room"])
    pdf_block("Color:", s["by_color"])
    pdf.output(SUMMARY_PDF)

    # Excel
    wb = Workbook(); ws = wb.active; ws.title = "Summary"
    ws.append(["Section","Name","Count"])
    for name, counter in [("Campus", s["by_campus"]), ("Category", s["by_category"]),
                          ("Status", s["by_status"]), ("Room", s["by_room"]), ("Color", s["by_color"])]:
        for k, v in counter.items():
            ws.append([name, k, v])
    ws2 = wb.create_sheet("Current View")
    ws2.append(HEADERS)
    for r in current_view_rows if current_view_rows else read_all_rows():
        ws2.append(r)
    wb.save(SUMMARY_XLSX)

    info_var.set("Özet kaydedildi: inventory_summary.pdf / inventory_summary.xlsx")

# ---------- Charts ----------
def show_charts():
    """Show charts with error handling"""
    try:
        # Set the backend explicitly
        import matplotlib
        matplotlib.use('TkAgg')
        import matplotlib.pyplot as plt
        
        s = make_summary()
        
        # Create a figure with subplots
        fig, axes = plt.subplots(2, 2, figsize=(12, 10))
        fig.suptitle('Avrasya Üniversitesi - Demirbaş İstatistikleri', fontsize=16, fontweight='bold')
        
        # Campus chart
        if s["by_campus"]:
            items, vals = zip(*s["by_campus"].most_common())
            axes[0,0].bar(items, vals, color='skyblue')
            axes[0,0].set_title('Yerleşke Bazlı Dağılım')
            axes[0,0].tick_params(axis='x', rotation=30)
        
        # Category chart
        if s["by_category"]:
            items, vals = zip(*s["by_category"].most_common())
            axes[0,1].bar(items, vals, color='lightgreen')
            axes[0,1].set_title('Kategori Bazlı Dağılım')
            axes[0,1].tick_params(axis='x', rotation=30)
        
        # Status chart
        if s["by_status"]:
            items, vals = zip(*s["by_status"].most_common())
            axes[1,0].bar(items, vals, color='lightcoral')
            axes[1,0].set_title('Durum Bazlı Dağılım')
            axes[1,0].tick_params(axis='x', rotation=0)
        
        # Color chart
        if s["by_room"]:
            items, vals = zip(*s["by_room"].most_common())
            axes[1,1].bar(items, vals, color='gold')
            axes[1,1].set_title('Oda Bazlı Dağılım')
            axes[1,1].tick_params(axis='x', rotation=30)
        
        plt.tight_layout()
        plt.show()
        
    except Exception as e:
        error_msg = f"Grafik gösterilirken hata: {str(e)}"
        print(f"Chart Error: {error_msg}")
        info_var.set(error_msg)
        messagebox.showerror("Grafik Hatası", error_msg)

# ---------- Backup ----------
def backup_excel():
    name = f"backup_inventory_{strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copyfile(EXCEL_FILE, name)
    info_var.set(f"Yedekleme oluşturuldu: {name}")

# ---------- Clock ----------
def tick():
    time_label.configure(text=now_time())
    time_label.after(1000, tick)

def export_room_report():
    room_name = filter_room_var.get().strip()
    if not room_name:
        info_var.set("Lütfen önce oda adını filtreye girin.")
        return

    rows = [r for r in read_all_rows() if str(r[6]).lower() == room_name.lower()]
    if not rows:
        info_var.set("Bu oda için kayıt bulunamadı.")
        return

    campus = rows[0][4]
    floor = rows[0][5]

    pdf = FPDF()
    pdf.add_page()

    pdf.set_font("Arial", "B", 16)
    pdf.cell(200, 12, "Avrasya University - Inventory Report", ln=True, align="C")
    pdf.ln(5)

    pdf.set_font("Arial", "", 12)
    pdf.set_fill_color(230, 230, 230)
    pdf.cell(200, 10, f"Campus: {campus}   |   Floor: {floor}   |   Room: {room_name}", ln=True, align="C", fill=True)
    pdf.ln(10)

    headers = ["Code", "Internal Code", "Inventory", "Category", "Status", "Date", "Color", "Turkish Name"]
    col_widths = [20, 20, 50, 25, 20, 25, 20, 40]

    pdf.set_font("Arial", "B", 9)
    pdf.set_fill_color(200, 200, 200)
    for h, wth in zip(headers, col_widths):
        pdf.cell(wth, 10, h, 1, 0, "C", fill=True)
    pdf.ln()

    fill = False
    pdf.set_font("Arial", "", 8)
    for r in rows:
        vals = [r[1], r[2], r[3], r[7], r[8], r[9], 
                r[10] if len(r) > 10 else "", 
                r[11] if len(r) > 11 else ""]
        for i, (v, wth) in enumerate(zip(vals, col_widths)):
            if i == 2 or i == 7:
                pdf.set_font("Arial", "B", 8)
            else:
                pdf.set_font("Arial", "", 8)
            pdf.cell(wth, 8, str(v), 1, 0, "C", fill=fill)
        pdf.ln()
        fill = not fill

    pdf.ln(5)
    pdf.set_font("Arial", "B", 12)
    pdf.cell(200, 10, f"Total items: {len(rows)}", ln=True, align="R")

    filename = f"room_report_{room_name}.pdf"
    pdf.output(filename)
    info_var.set(f"Oda raporu oluşturuldu: {filename}")

def clear_fields():
    code_var.set("")
    internal_code_var.set("")
    inventory_var.set("")
    campus_var.set("")
    floor_var.set("")
    room_var.set("")
    category_var.set("")
    status_var.set("")
    date_var.set("")
    color_var.set("")
    turkish_name_var.set("")
    info_var.set("Alanlar temizlendi.")

def show_about():
    about_win = CTkToplevel(w)
    about_win.title("About")
    about_win.geometry("700x600")

    txt = CTkTextbox(about_win, wrap="word", font=("Arial", 12))
    txt.pack(expand=True, fill="both", padx=10, pady=10)

    about_text = """
PROGRAM HAKKINDA

Avrasya Üniversitesi Demirbaş Takip Sistemi, üniversitedeki tüm demirbaş ve Demirbaş varlıklarının dijital olarak kaydedilmesi, yönetilmesi, sınıflandırılması, raporlanması ve izlenmesi için geliştirilmiş kapsamlı bir yazılımdır.

Bu yazılım yalnızca Avrasya Üniversitesi için üretilmiştir.
Üniversite dışındaki kullanım yasal işlem gerektirir.

PROGRAMIN AMAÇLARI

Tüm Demirbaşlerin merkezi bir sistemde toplanması

Hızlı, güvenilir ve doğru Demirbaş takibi

Detaylı raporlama ve analiz imkânı

Barkod ve QR kod destekli fiziksel takip

Zaman ve iş gücü tasarrufu

TEMEL ÖZELLİKLER
1. KAYIT YÖNETİMİ

Yeni kayıt ekleme

Mevcut kayıtları düzenleme

Kayıt silme veya arşivleme

Otomatik alan temizleme

2. EXCEL ENTEGRASYONU

Excel dosyalarından otomatik veri alma

Birden fazla sayfa desteği

Akıllı sütun tanıma (Adı, Cinsi, Renk, Adet vb.)

Türkçe karakter uyumluluğu

3. GELİŞMİŞ FİLTRELEME

Koda, iç koda, kategoriye, duruma göre filtreleme

Yerleşke, kat ve oda bazlı filtreleme

Renk ve Türkçe isim filtreleme

4. BARKOD & QR KOD SİSTEMİ

Standart Barkod: Sadece Demirbaş numarası

Detaylı Barkod: Numara + kısa isim + yer bilgisi

QR Kod: Ürüne ait tüm bilgileri içeren detaylı kayıt

5. RAPORLAMA

CSV olarak dışa aktarma

Profesyonel PDF raporları

Oda bazlı detaylı raporlar

İstatistiksel özet raporlar

6. GÖRSELLEŞTİRME

Yerleşke bazlı grafikler

Kategori dağılımı grafikleri

Durum analizi grafikler

Renk bazlı istatistikler

ARAYÜZ KULLANIMI
SOL PANEL – KAYIT FORMU

Aşağıdaki alanları doldurarak yeni kayıt oluşturabilirsiniz:

Kod: Otomatik oluşturulan benzersiz kod

Dahili Kod: İç referans kodu

Demirbaş Öğesi: Ürün açıklaması (İngilizce)

Türkçe İsim: Ürünün Türkçe adı (zorunlu)

Renk: Ürün rengi

Kategori: Ön tanımlı kategori

Status: Yeni / Kullanılmış / Hasarlı

Campus: Yerleşke (zorunlu)

Floor: Kat

Room: Oda

Date: Kayıt tarihi

MERKEZ PANEL – İŞLEMLER
Temel İşlemler

Kaydetmek: Yeni kayıt ekler

Düzenlemek: Seçili kaydı forma yükler

Kaydet Düzenlendi: Düzenlenen kaydı günceller

Silmek: Seçili kaydı siler

Alanları Temizle: Formu sıfırlar

Raporlama

CSV’ye Aktar

PDF’ye Aktar

Özeti Göster

Oda Raporu (PDF)

Barkod İşlemleri

Barkod Göster

QR Kod Göster

Detaylı Barkod

Tüm Barkodlar / Tüm QR Kodlar

Sistem İşlemleri

Excel Import

Excel Yedekle

Grafikleri Göster

Özeti Kaydet (PDF + Excel)

SAĞ PANEL – FİLTRELER

Koda göre

Dahili koda göre

Kampüse göre

Odaya göre

Demirbaş öğesine göre

Türkçe isme göre

Renge göre

BAŞLANGIÇ REHBERİ
ADIM 1 – PROGRAMI AÇMA

Program başlatılır

inventory.xlsx otomatik olarak oluşur

Var olan tüm veriler yüklenir

ADIM 2 – YENİ KAYIT EKLEME

Sol paneldeki form doldurulur

“Demirbaş Öğesi” ve “Campus” zorunlu

“Kaydetmek” butonuna basılır

Kayıt tabloya eklenir

ADIM 3 – EXCEL’DEN VERİ AKTARMA

“Excel Import” butonuna tıklanır

Pelitli_Demirbaş-idarikat.xlsx seçilir

Program sütun adlarını otomatik tanır

Tüm sayfa ve kayıtlar içe aktarılır

ADIM 4 – KAYIT DÜZENLEME

Tablodan kayıt seçilir

“Düzenlemek” butonuna basılır

Form otomatik olarak dolar

Düzenleme yapıp “Kaydet Düzenlendi” butonuna tıklanır

ADIM 5 – BARKOD OLUŞTURMA

Kayıt seçilir

Kullanım amacına göre barkod türü seçilir

Hızlı takip → Standart Barkod

Detaylı bilgi → QR Kod

Orta seviye → Detaylı Barkod

Oluşturulan kod yazdırılıp malzemeye yapıştırılır

RAPORLAMA ÖZELLİKLERİ
1. Oda Raporu

Seçili odadaki tüm Demirbaş

Kategori ve durum analizleri

PDF formatında çıktı

2. İstatistiksel Özet

Yerleşke dağılımı

Kategori analizi

Durum raporu

Oda ve renk istatistikleri

3. Grafikler

Çubuk grafikler

Pasta grafikler

Trend analizleri

BARKOD SİSTEMİ DETAYLARI
Standart Barkod

İçerik: Sadece Demirbaş numarası

Kullanım: Hızlı takip

Detaylı Barkod

İçerik: Numara + kısa isim + yer

Kullanım: Orta seviye bilgi ihtiyaçları

QR Kod

Tüm bilgiler:

Demirbaş kodu

İsimler

Yerleşke, kat, oda

Kategori, durum, renk

Kayıt tarihi, ID

Kullanım: En detaylı kayıt türü

YEDEKLEME ve GÜVENLİK
Otomatik Yedekleme

Her işlemde otomatik kaydetme

Excel formatında saklama

Manuel Yedekleme

“Excel’i Yedekle” butonuna tıklayın

Zaman damgalı yedek oluşturulur

SIK KARŞILAŞILAN SORUNLAR
✔ Excel içe aktarılmıyor

Dosya .xlsx olmalı

Sütunlar Türkçe olmalı

Dosya bozuk olmamalı

✔ Barkod çıkmıyor

Gereken kütüphaneleri yükleyin:

pip install python-barcode[qrcode] pillow qrcode[pil]


Programı yeniden başlatın

✔ Türkçe karakter sorunu

Program otomatik Unicode destekler

PDF çıktılarında Türkçe uyumludur

✔ Veri kaybı

Düzenli yedek alın

Otomatik kaydetme aktiftir

GELİŞTİRİCİ ve DESTEK

Geliştirici: Ramin Entezar

Kurum: Avrasya Üniversitesi

Yıl: 2025

Sürüm: 2.0 – Barkod Destekli

Destek için:
Programdaki About bölümünü inceleyin, hata mesajlarını not alın ve teknik destek ile iletişime geçin.

EN İYİ UYGULAMA ÖNERİLERİ
Veri Girişi

Zorunlu alanları eksiksiz doldurun

Tutarlı kategori ve durum seçin

Anlamlı Türkçe isim kullanın

Excel Aktarımında

Standart şablon kullanın

Sayfa isimlerini oda adı olarak kaydedin

Aktarım öncesi yedek alın

Barkod Kullanımı

Değerli malzemelere QR kod ekleyin

Hızlı tarama için standart barkod kullanın

Kodları nem ve darbeye karşı koruyun

Filtreleme

Büyük verilerde oda raporu kullanın

Kategori ve durum filtreleriyle analiz yapın  




Programcı: Ramin Entezar
 """

    txt.insert("1.0", about_text)
    txt.configure(state="disabled")

# ================= UI =================
init_workbook()

set_appearance_mode("light")
set_default_color_theme("blue")

w = CTk()
w.geometry("1800x1000")
w.title("Avrasya Üniversitesi Demirbaş Takip Sistemi - V3.0 (Kategori Mevcut)")

# ========== اضافه کردن لوگو به عنوان پنجره ==========
try:
    w.iconbitmap(LOGO_FILE)  # برای ویندوز
    # w.iconphoto(True, ImageTk.PhotoImage(Image.open(LOGO_FILE)))  # برای لینوکس/مک
except:
    print("⚠️ لوگو یافت نشد. برنامه بدون آیکون اجرا خواهد شد.")

# Variables
campus_var = StringVar(); floor_var = StringVar(); room_var = StringVar()
code_var = StringVar(); internal_code_var = StringVar(); inventory_var = StringVar()
category_var = StringVar(); status_var = StringVar(); date_var = StringVar()
color_var = StringVar(); turkish_name_var = StringVar()

filter_code_var = StringVar()
filter_internal_code_var = StringVar()
filter_campus_var = StringVar(); filter_room_var = StringVar(); filter_inventory_var = StringVar()
filter_turkish_name_var = StringVar(); filter_color_var = StringVar(); filter_category_var = StringVar()

result_count_var = StringVar(value="Gösterim Sayısı: 0")
info_var = StringVar()

# ========== ایجاد هدر با لوگو ==========
# فریم برای هدر
header_frame = CTkFrame(w, width=1600, height=80, corner_radius=10, fg_color="#2962FF")  # رنگ آبی تیره
header_frame.place(x=150, y=10)

try:
    # بارگذاری لوگو برای هدر با استفاده از CTkImage
    header_logo_img = Image.open(LOGO_FILE)
    header_logo_img = header_logo_img.resize((60, 60), Image.LANCZOS)
    
    # تبدیل به CTkImage
    header_ctk_image = CTkImage(light_image=header_logo_img, dark_image=header_logo_img, size=(60, 60))
    
    # نمایش لوگو در هدر
    header_logo_label = CTkLabel(header_frame, image=header_ctk_image, text="", fg_color="transparent")
    header_logo_label.place(x=620, y=10)
except Exception as e:
    print(f"خطا در بارگذاری لوگو: {e}")
    # اگر لوگو پیدا نشد، از متن استفاده کن
    header_logo_label = CTkLabel(header_frame, text="AURASIA", font=("Arial", 20, "bold"), 
                                 text_color="white", fg_color="transparent")
    header_logo_label.place(x=20, y=25)

# عنوان برنامه در هدر
CTkLabel(header_frame, text="AVRASYA ÜNİVERSİTESİ", 
         font=("Arial", 24, "bold"), text_color="white", fg_color="transparent").place(x=700, y=15)
CTkLabel(header_frame, text="Demirbaş Takip Sistemi", 
         font=("Arial", 20, "bold"), text_color="#CCCCCC", fg_color="transparent").place(x=730, y=45)



# Main frames (موقعیت‌ها کمی پایین‌تر آمده‌اند)
left_frame = CTkFrame(w, width=400, height=500, corner_radius=10)
left_frame.place(x=150, y=100)  # y از 100 شروع می‌شود

center_frame = CTkFrame(w, width=350, height=550, corner_radius=10)
center_frame.place(x=625, y=100)  # y از 100 شروع می‌شود

right_frame = CTkFrame(w, width=450, height=500, corner_radius=10)
right_frame.place(x=1050, y=100)  # y از 100 شروع می‌شود

table_frame = CTkFrame(w, width=1800, height=400, corner_radius=10)
table_frame.place(x=20, y=680)  # y از 680 شروع می‌شود

# Left form (بدون تغییر)
CTkLabel(left_frame, text='Kayitlar', font=("Arial", 18, "bold"),fg_color="#A8BEF8" ,corner_radius=5).place(x=130, y=10)

CTkLabel(left_frame, text='Kod', font=("Arial", 12, "bold")).place(x=20, y=60)
CTkEntry(left_frame, textvariable=code_var, width=220).place(x=150, y=60)

CTkLabel(left_frame, text='Dahili Kod', font=("Arial", 12, "bold")).place(x=20, y=100)
CTkEntry(left_frame, textvariable=internal_code_var, width=220).place(x=150, y=100)

CTkLabel(left_frame, text='Demırbaş Öğesi', font=("Arial", 12, "bold")).place(x=20, y=140)
CTkEntry(left_frame, textvariable=inventory_var, width=220).place(x=150, y=140)

CTkLabel(left_frame, text='Kullanan Personel', font=("Arial", 12, "bold")).place(x=20, y=180)
CTkEntry(left_frame, textvariable=turkish_name_var, width=220).place(x=150, y=180)
CTkLabel(left_frame, text='Renk', font=("Arial", 12, "bold")).place(x=20, y=220)
color_values = ('Siyah', 'Beyaz', 'Kahverengi', 'Gri', 'Mavi', 'Kırmızı', 'Yeşil', 'Turkuaz', 'Bej', 'Diğer')
color_cb = CTkComboBox(left_frame, variable=color_var, width=220, values=color_values)
color_cb.place(x=150, y=220)
CTkLabel(left_frame, text='Kategori', font=("Arial", 12, "bold")).place(x=20, y=260)
category_values = ('Mobilya ve Ofis Esyalari','Egitim Araclari','Elektronik Cihazlar',
                   'Laboratuvar Ekipmanlari','Atolye Araclari','Hizmet ve Konfor urunleri',
                   'Spor Ekipmanlari','Araclar','Altyapi Ekipmanlari')
category_cb = CTkComboBox(left_frame, variable=category_var, width=220, values=category_values)
category_cb.place(x=150, y=260)

CTkLabel(left_frame, text='Durum', font=("Arial", 12, "bold")).place(x=20, y=300)
status_values = ('Yeni','Kullanılmış','Hasarlı')
status_cb = CTkComboBox(left_frame, variable=status_var, width=220, values=status_values)
status_cb.place(x=150, y=300)

CTkLabel(left_frame, text='Kampüs', font=("Arial", 12, "bold")).place(x=20, y=340)
campus_values = ('Pelitli', 'omer Yildiz', 'Yomra', 'Kasustu')
campus_cb = CTkComboBox(left_frame, variable=campus_var, width=220, values=campus_values)
campus_cb.place(x=150, y=340)

CTkLabel(left_frame, text='Kat', font=("Arial", 12, "bold")).place(x=20, y=380)
floor_values = ('Kat -1','Kat 0','Kat 1','Kat 2','Kat 3','Kat 4')
floor_cb = CTkComboBox(left_frame, variable=floor_var, width=220, values=floor_values)
floor_cb.place(x=150, y=380)

CTkLabel(left_frame, text='Oda', font=("Arial", 12, "bold")).place(x=20, y=420)
room_values = ('YAZI İŞLERİ', 'REKTÖRLÜK', 'REKTÖRLÜK DİNLENME ODASI', 'REKTÖR ÖZEL KALEM', 
               'GENEL SEKRETERLİK', 'GENEL SEKRETERLİK DİNLENME ODASI', 'BAŞKAN DANIŞMA ODASI',
               'MÜTEVELLİ HEYET ÖZEL KALEM', 'MÜTEVELLİ HEYET BAŞKANI', 'BAŞKAN TOPLANTI ODASI',
               'BAŞKAN DİNLENME ODASI', '(İDARİ KAT)KORİDOR')
room_cb = CTkComboBox(left_frame, variable=room_var, width=220, values=room_values)
room_cb.place(x=150, y=420)

CTkLabel(left_frame, text='Tarih', font=("Arial", 12, "bold")).place(x=20, y=460)
date_entry = DateEntry(left_frame, textvariable=date_var, width=24, background='darkblue',
                       foreground='white', borderwidth=2, date_pattern='yyyy-mm-dd')
date_entry.place(x=150, y=460)

# Center (بدون تغییر)
CTkLabel(center_frame, text='İşlemler', font=("Arial", 18, "bold"),fg_color="#A8BEF8" ,corner_radius=5).place(x=130, y=10)

CTkButton(center_frame, text='Yeni Kayıt', command=save_record, width=100 ,fg_color="#2962FF", hover_color="#0039CB",font=("Arial", 14, "bold")).place(x=60, y=70)
CTkButton(center_frame, text='Düzenle', command=edit_selected, width=100,fg_color="#2962FF", hover_color="#0039CB" ,font=("Arial", 14, "bold")).place(x=190, y=70)
CTkButton(center_frame, text='Kaydet', command=save_edited, width=100,fg_color="green", hover_color="darkgreen",font=("Arial", 14, "bold")).place(x=190, y=110)
CTkButton(center_frame, text='Sil', command=delete_selected, width=100 , fg_color="#E53935", hover_color="#C62828",font=("Arial", 14, "bold")).place(x=60, y=110)
CTkButton(center_frame, text='Temizle', command=clear_fields, width=100, fg_color="#2962FF", hover_color="#0039CB",font=("Arial", 14, "bold")).place(x=60, y=150)

CTkButton(center_frame, text='Özet', command=show_summary, width=100, fg_color="#2962FF", hover_color="#0039CB",font=("Arial", 14, "bold")).place(x=190, y=150)

CTkButton(center_frame, text='Grafikleri Göster', command=show_charts, width=120,
          fg_color="#2962FF", hover_color="#0039CB",font=("Arial", 14, "bold")).place(x=30, y=190)
CTkButton(center_frame, text="Excel'i Yedekle", command=backup_excel, width=120,
          fg_color="#2962FF", hover_color="#0039CB",font=("Arial", 14, "bold")).place(x=180, y=190)

CTkButton(center_frame, text="Özeti Kaydet", command=save_summary_files, width=120,
          fg_color="#2962FF", hover_color="#0039CB",font=("Arial", 14, "bold")).place(x=40, y=230)
CTkButton(center_frame, text="Excel Import", command=simple_import_excel, width=120, 
          fg_color="green", hover_color="darkorange",font=("Arial", 14, "bold")).place(x=180, y=230)

# NEW BARCODE BUTTONS
CTkLabel(center_frame, text='QR Kod İşlemleri', font=("Arial", 14, "bold")).place(x=110, y=270)

# QR Code buttons
CTkButton(center_frame, text="QR Kod Göster", command=show_qrcode, width=120,
          fg_color="#2962FF", hover_color="#0039CB",font=("Arial", 14, "bold")).place(x=40, y=310)

CTkButton(center_frame, text="Tüm QR Kodlar", command=generate_all_qrcodes, width=120,
          fg_color="#2962FF", hover_color="#0039CB",font=("Arial", 14, "bold")).place(x=180, y=310)

# IMAGE UPLOAD BUTTONS - NEW SECTION
CTkLabel(center_frame, text='Görsel İşlemleri', font=("Arial", 14, "bold")).place(x=110, y=350)

CTkButton(center_frame, text="Görsel Yükle", command=upload_item_image, width=120,
          fg_color="orange", hover_color="darkorange",font=("Arial", 14, "bold")).place(x=40, y=390)

CTkButton(center_frame, text="Görseli Göster", command=show_item_image, width=120,
          fg_color="orange", hover_color="darkorange",font=("Arial", 14, "bold")).place(x=180, y=390)

# ROOM QR CODE BUTTONS - NEW SECTION
CTkLabel(center_frame, text='Oda QR İşlemleri', font=("Arial", 14, "bold")).place(x=110, y=430)

CTkButton(center_frame, text="Oda QR Göster", command=show_room_qrcode, width=120,
          fg_color="#4CAF50", hover_color="#388E3C",font=("Arial", 14, "bold")).place(x=40, y=470)

CTkButton(center_frame, text="Tüm Oda QR'ları", command=generate_all_room_qrcodes, width=120,
          fg_color="#4CAF50", hover_color="#388E3C",font=("Arial", 14, "bold")).place(x=180, y=470)

CTkButton(center_frame, text="Oda Grup QR", command=show_grouped_item_qrcodes, width=120,
          fg_color="#4CAF50", hover_color="#388E3C",font=("Arial", 14, "bold")).place(x=110, y=510)

# Right frame (بدون تغییر)
CTkLabel(right_frame, text='Filtreler', font=("Arial", 18, "bold") ,fg_color="#A8BEF8" ,corner_radius=5).place(x=180, y=10)

CTkLabel(right_frame, text='Koda Göre', font=("Arial", 12, "bold")).place(x=20, y=60)
CTkEntry(right_frame, textvariable=filter_code_var, width=200).place(x=180, y=60)

CTkLabel(right_frame, text="Dahili Koda Göre", font=("Arial", 12, "bold")).place(x=20, y=100)
CTkEntry(right_frame, textvariable=filter_internal_code_var, width=200).place(x=180, y=100)

CTkLabel(right_frame, text='Kampüse Göre', font=("Arial", 12, "bold")).place(x=20, y=140)
CTkEntry(right_frame, textvariable=filter_campus_var, width=200).place(x=180, y=140)

CTkLabel(right_frame, text='Odaya Göre', font=("Arial", 12, "bold")).place(x=20, y=180)
CTkEntry(right_frame, textvariable=filter_room_var, width=200).place(x=180, y=180)

CTkLabel(right_frame, text='Demirbaşa Göre', font=("Arial", 12, "bold")).place(x=20, y=220)
CTkEntry(right_frame, textvariable=filter_inventory_var, width=200).place(x=180, y=220)

CTkLabel(right_frame, text='Kullanan Personel Göre', font=("Arial", 12, "bold")).place(x=20, y=260)
CTkEntry(right_frame, textvariable=filter_turkish_name_var, width=200).place(x=180, y=260)

CTkLabel(right_frame, text='Kategoriye Göre', font=("Arial", 12, "bold")).place(x=20, y=300)
CTkEntry(right_frame, textvariable=filter_category_var, width=200).place(x=180, y=300)

# CTkLabel(right_frame, text='Renge Göre Filtrele').place(x=20, y=290)
# CTkEntry(right_frame, textvariable=filter_color_var, width=200).place(x=180, y=290)

CTkButton(right_frame, text='Filtrele', command=filter_view, width=120,
          fg_color="orange", hover_color="darkorange" , font=("Arial", 14, "bold")).place(x=200, y=380)
CTkButton(right_frame, text='Tümünü Göster', command=show_all, width=120,
          fg_color="#2962FF", hover_color="#0039CB" , font=("Arial", 14, "bold")).place(x=200, y=420)

# Table
cols = ("ID","Kod","Dahili Kod","Demirbaş Öğesi","Kampüs","Kat","Oda","Kategori","Durum","Tarih","Renk","Kullanan Personel","Görsel")
tree = ttk.Treeview(table_frame, columns=cols, show="headings")
for c in cols:
    tree.heading(c, text=c)
    if c == "ID":
        tree.column(c, width=0, stretch=False)
    elif c == "Görsel":
        tree.column(c, width=40, stretch=False)  # Small column for image indicator
    elif c in ["Color", "Turkish Name"]:
        tree.column(c, width=120, stretch=True)
    else:
        tree.column(c, width=100, stretch=True)
tree.place(x=10, y=10, width=1770, height=340)
scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
tree.configure(yscroll=scrollbar.set)
scrollbar.place(x=1782, y=12, height=335)

# ========== اضافه کردن لوگو در قسمت اطلاعات سمت راست ==========
# فریم برای لوگو و اطلاعات
logo_info_frame = CTkFrame(w, width=100, height=200, corner_radius=10)
logo_info_frame.place(x=1570, y=100)

try:
    # بارگذاری لوگو برای قسمت اطلاعات با استفاده از CTkImage
    small_logo_img = Image.open(LOGO_FILE)
    small_logo_img = small_logo_img.resize((80, 80), Image.LANCZOS)
    
    # تبدیل به CTkImage
    small_ctk_image = CTkImage(light_image=small_logo_img, dark_image=small_logo_img, size=(80, 80))
    
    # نمایش لوگو
    small_logo_label = CTkLabel(logo_info_frame, image=small_ctk_image, text="", fg_color="transparent")
    small_logo_label.pack(pady=10)
except Exception as e:
    print(f"خطا در بارگذاری لوگو کوچک: {e}")
    # اگر لوگو پیدا نشد
    CTkLabel(logo_info_frame, text="AVRASYA", font=("Arial", 16, "bold")).pack(pady=10)

# اطلاعات در زیر لوگو
CTkLabel(logo_info_frame, text="  AVRASYA ÜNİVERSİTESİ  ", 
         font=("Metropolis", 14, "bold")).pack(pady=5)
CTkLabel(logo_info_frame, text="Demirbaş Takip Sistemi", 
         font=("Metropolis", 12, "bold")).pack(pady=2)
CTkLabel(logo_info_frame, text="Bu program \nAvrasya Üniversitesi \n tarafından tasarlanmıştır.", 
         font=("Metropolis", 10, )).pack(pady=2)
CTkLabel(logo_info_frame, text="© 2025", 
         font=("Metropolis", 9)).pack(pady=2)

# دکمه About
CTkButton(logo_info_frame, text='Hakkında', command=show_about, width=140,
          fg_color="#2962FF", hover_color="#0039CB").pack(pady=10)

# Clock and info (موقعیت تغییر کرده)
time_label = CTkLabel(w, font=("Arial", 12, "bold"),width=180, text_color="black",fg_color="azure3",corner_radius=5)
time_label.place(x=1570, y=420)  # موقعیت جدید
tick()

CTkLabel(w, textvariable=result_count_var, font=("Arial", 12, "bold"), 
         fg_color="spring green", text_color="black", width=180, corner_radius=5).place(x=1570, y=470)

CTkLabel(w, textvariable=info_var, fg_color="azure3", text_color="blue", 
         font=("Arial", 12, "bold"), corner_radius=5).place(x=1570, y=570)

# Start
load_to_tree()
refresh_dynamic_choices()
w.mainloop()

print("Initializing barcode system...")
print(f"Barcode directory: {os.path.abspath(BARCODE_DIR)}")
print(f"Directory exists: {os.path.exists(BARCODE_DIR)}")