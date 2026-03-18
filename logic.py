import os
import pandas as pd
from openpyxl import load_workbook, Workbook
from uuid import uuid4
from time import strftime
import qrcode
import barcode
from barcode.writer import ImageWriter
from fpdf import FPDF
import streamlit as st
from PIL import Image

def tr_to_eng(text):
    if not isinstance(text, str):
        return str(text)
    replacements = {
        'ı': 'i', 'İ': 'I',
        'ş': 's', 'Ş': 'S',
        'ğ': 'g', 'Ğ': 'G',
        'ü': 'u', 'Ü': 'U',
        'ö': 'o', 'Ö': 'O',
        'ç': 'c', 'Ç': 'C'
    }
    for search, replace in replacements.items():
        text = text.replace(search, replace)
    return text

# --- Configuration ---
EXCEL_FILE = "inventory.xlsx"
BARCODE_DIR = "barcodes"
QRCODE_DIR = "qrcodes"
IMAGES_DIR = "images"
LOGO_FILE = "LOGO1.png"
HEADERS = ["ID", "Code", "Internal_Code", "Inventory_Item", "Campus", "Floor", "Room", "Category", "Status", "Date", "Color", "Turkish_Name", "Image_Path"]

# --- Initialization ---
def init_directories():
    os.makedirs(BARCODE_DIR, exist_ok=True)
    os.makedirs(QRCODE_DIR, exist_ok=True)
    os.makedirs(IMAGES_DIR, exist_ok=True)

def init_db():
    init_directories()
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.append(HEADERS)
        wb.save(EXCEL_FILE)
    else:
        # Compatibility check
        df = pd.read_excel(EXCEL_FILE)
        if list(df.columns) != HEADERS:
            # Simple header fix if needed, but following the original logic
            pass 

def get_file_last_modified():
    return os.path.getmtime(EXCEL_FILE) if os.path.exists(EXCEL_FILE) else 0

@st.cache_data
def _load_data_cached(last_modified):
    if not os.path.exists(EXCEL_FILE):
        init_db()
    # Read with explicit column names to ensure consistency
    try:
        return pd.read_excel(EXCEL_FILE)
    except Exception as e:
        st.error(f"Excel read error: {e}")
        return pd.DataFrame(columns=HEADERS)

def load_data():
    return _load_data_cached(get_file_last_modified())

def clear_cache():
    st.cache_data.clear()

def save_item(item_data):
    """Save or update an item in the Excel file"""
    df = load_data()
    
    if "ID" not in item_data or not item_data["ID"]:
        item_data["ID"] = str(uuid4())
        item_data["Date"] = strftime('%Y-%m-%d')
        new_row = pd.DataFrame([item_data])
        df = pd.concat([df, new_row], ignore_index=True)
    else:
        idx = df[df['ID'] == item_data['ID']].index
        if not idx.empty:
            for col in item_data:
                df.at[idx[0], col] = item_data[col]
        else:
            new_row = pd.DataFrame([item_data])
            df = pd.concat([df, new_row], ignore_index=True)
            
    df.to_excel(EXCEL_FILE, index=False)
    return item_data["ID"]

def delete_item(item_id):
    df = load_data()
    df = df[df['ID'] != item_id]
    df.to_excel(EXCEL_FILE, index=False)

# --- Asset Generation ---
def generate_qrcode(item_data):
    item_code = item_data.get('Code', 'N/A')
    qr = qrcode.QRCode(version=2, box_size=8, border=4)
    
    qr_content = f""" AVRASYA ÜNİVERSİTESİ - DEMİRBAŞ SİSTEMİ 

TEMEL BİLGİLER:
├─ Demirbaş Numarası: {item_code}
├─ Dahili Kod: {item_data.get('Internal_Code', '')}
├─ Kullanan Personel: {item_data.get('Turkish_Name', '')}
├─ Ad: {item_data.get('Inventory_Item', '')}

YERLEŞİM BİLGİSİ:
├─ Yerleşke: {item_data.get('Campus', '')}
├─ Oda: {item_data.get('Room', '')}

KATEGORİ ve DURUM:
├─ Kategori: {item_data.get('Category', '')}
├─ Durum: {item_data.get('Status', '')}
├─ Renk: {item_data.get('Color', 'Belirtilmemiş')}

KAYIT BİLGİSİ:
├─ Kayıt Tarihi: {item_data.get('Date', '')}
├─ Benzersiz ID: {item_data.get('ID', '')}

AÇIKLAMA:
Avrasya Üniversitesi demirbaş sistemine kayıtlı demirbaş malzemedir.
Lütfen bu malzemenin yerini değiştirmeyiniz veya başka birime taşımayınız.

 İLETİŞİM:
Demirbaş Yönetimi - Avrasya Üniversitesi
"""
    qr.add_data(qr_content)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    path = os.path.join(QRCODE_DIR, f"qrcode_{item_code}.png")
    img.save(path)
    return path

def generate_all_qrcodes():
    """Tüm demirbaşlar için QR kodlarını qrcodes/ klasörüne üretir"""
    df = load_data()
    generated_count = 0
    for _, item in df.iterrows():
        item_dict = item.to_dict()
        if generate_qrcode(item_dict):
            generated_count += 1
    return generated_count

def generate_barcode(item_code):
    try:
        code128 = barcode.get_barcode_class('code128')
        it_code = str(item_code).strip()
        bar = code128(it_code, writer=ImageWriter())
        path = os.path.join(BARCODE_DIR, f"barcode_{it_code}")
        saved_path = bar.save(path)
        return saved_path
    except Exception as e:
        print(f"Barcode error: {e}")
        return None

# --- PDF Generation ---
def create_item_pdf(item_data, qr_path=None):
    item_code = item_data.get('Code', 'N/A')
    pdf = FPDF()
    pdf.add_page()
    
    # Logo
    if os.path.exists(LOGO_FILE):
        pdf.image(LOGO_FILE, x=75, y=10, w=60)
        pdf.set_y(40)
    else:
        pdf.set_y(10)
        
    pdf.set_font("Arial", 'B', 16)
    pdf.cell(200, 10, tr_to_eng("AVRASYA UNIVERSITESI - DEMIRBAS SISTEMI"), ln=True, align='C')
    pdf.ln(10)
    
    pdf.set_font("Arial", '', 12)
    for col in HEADERS:
        if col != "Image_Path":
            val = tr_to_eng(f"{col}: {item_data.get(col, '')}")
            pdf.cell(200, 8, val, ln=True)
            
    if qr_path and os.path.exists(qr_path):
        pdf.ln(10)
        pdf.image(qr_path, x=70, y=pdf.get_y(), w=70)
        
    filename = f"Demirbas_{tr_to_eng(str(item_code)).replace(' ', '_')}.pdf"
    pdf.output(filename)
    return filename

def create_room_pdf(room_name, items):
    pdf = FPDF()
    pdf.add_page()
    
    if os.path.exists(LOGO_FILE):
        pdf.image(LOGO_FILE, x=85, y=10, w=40)
        pdf.set_y(35)
    else:
        pdf.set_y(10)
        
    pdf.set_font("Arial", 'B', 18)
    pdf.cell(200, 10, "AVRASYA UNIVERSITESI", ln=True, align='C')
    pdf.set_font("Arial", 'B', 14)
    room_eng = tr_to_eng(str(room_name))
    pdf.cell(200, 10, f"ODA ENVANTER LISTESI - {room_eng}", ln=True, align='C')
    pdf.ln(5)
    
    pdf.set_font("Arial", 'B', 10)
    pdf.cell(40, 10, "Kod", 1)
    pdf.cell(100, 10, "Malzeme", 1)
    pdf.cell(50, 10, "Durum", 1, ln=True)
    
    pdf.set_font("Arial", '', 9)
    for it in items:
        pdf.cell(40, 8, tr_to_eng(str(it.get('Code', ''))), 1)
        pdf.cell(100, 8, tr_to_eng(str(it.get('Turkish_Name', ''))), 1)
        pdf.cell(50, 8, tr_to_eng(str(it.get('Status', ''))), 1, ln=True)
        
    filename = f"Room_{room_eng.replace(' ', '_')}.pdf"
    pdf.output(filename)
    return filename

def import_from_excel(file):
    try:
        df_new = pd.read_excel(file)
        # Basic mapping check - in a real app, we'd do more robust matching
        # For now, we assume it matches the expected structure or has overlapping names
        df_current = load_data()
        
        # Ensure ID column exists for new items
        if "ID" not in df_new.columns:
            df_new["ID"] = [str(uuid4()) for _ in range(len(df_new))]
            
        # Merge or append
        df_combined = pd.concat([df_current, df_new], ignore_index=True).drop_duplicates(subset=["ID"], keep='last')
        df_combined.to_excel(EXCEL_FILE, index=False)
        return len(df_new)
    except Exception as e:
        print(f"Import error: {e}")
        return 0

# --- Image Handling ---
def save_image(uploaded_file, item_code):
    if uploaded_file is not None:
        extension = uploaded_file.name.split('.')[-1]
        filename = f"item_{item_code}.{extension}"
        path = os.path.join(IMAGES_DIR, filename)
        with open(path, "wb") as f:
            f.write(uploaded_file.getbuffer())
        return path
    return None
